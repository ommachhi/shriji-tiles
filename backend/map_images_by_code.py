from __future__ import annotations

from pathlib import Path
import re
import shutil

from openpyxl import load_workbook


def code_to_filename(code: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(code).strip().lower()).strip("-")
    if not slug:
        slug = "unknown"
    return f"{slug}.png"


def update_excel_images(excel_path: Path, images_dir: Path) -> tuple[int, int, Path]:
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_cells]
    index = {name: idx for idx, name in enumerate(headers)}

    code_col = index.get("code")
    image_col = index.get("image")
    image_file_col = index.get("image_file")

    if code_col is None or image_col is None or image_file_col is None:
        workbook.close()
        raise ValueError(f"Missing required columns in {excel_path.name}")

    updated_rows = 0
    copied_files = 0

    for row in sheet.iter_rows(min_row=2):
        code = row[code_col].value
        image_value = row[image_col].value

        if code is None or image_value is None:
            continue

        code_str = str(code).strip()
        image_str = str(image_value).strip()
        if not code_str or not image_str:
            continue

        source_name = Path(image_str).name
        source_path = images_dir / source_name
        if not source_path.exists() or not source_path.is_file():
            continue

        target_name = code_to_filename(code_str)
        target_path = images_dir / target_name

        if not target_path.exists():
            shutil.copy2(source_path, target_path)
            copied_files += 1

        row[image_col].value = f"/images/{target_name}"
        row[image_file_col].value = target_name
        updated_rows += 1

    saved_path = excel_path
    try:
        workbook.save(excel_path)
    except PermissionError:
        saved_path = excel_path.with_name(f"{excel_path.stem}_codeimg{excel_path.suffix}")
        workbook.save(saved_path)
    workbook.close()
    return updated_rows, copied_files, saved_path


def main() -> None:
    base = Path(__file__).resolve().parent
    images_dir = base / "images"
    excels = [
        base / "catalog_pages_4_43.xlsx",
        base / "kohler_pages_1_40.xlsx",
    ]

    total_rows = 0
    total_copies = 0

    for excel in excels:
        if not excel.exists():
            print(f"skip: {excel.name} (not found)")
            continue

        rows, copies, saved_path = update_excel_images(excel, images_dir)
        total_rows += rows
        total_copies += copies
        print(f"{excel.name}: updated_rows={rows}, copied_files={copies}, saved={saved_path.name}")

    print(f"done: total_updated_rows={total_rows}, total_copied_files={total_copies}")


if __name__ == "__main__":
    main()
