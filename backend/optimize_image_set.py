#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
IMAGES_DIR = BASE_DIR / "images"
CACHE_FILES = [BASE_DIR / "catalog_cache.json", BASE_DIR / "kohler_cache.json"]
EXCEL_FILES = [BASE_DIR / "aquant_catalog_full.xlsx", BASE_DIR / "kohler_catalog_full.xlsx"]
SUPPORTED = {".png", ".jpg", ".jpeg"}


@dataclass
class ImageInfo:
    rel: str
    path: Path
    size_bytes: int
    binary_key: str


def _safe_rel(path: Path) -> str:
    return str(path.relative_to(IMAGES_DIR)).replace("\\", "/")


def _sanitize_code(value: str) -> str:
    code = re.sub(r"\s*([+\-/])\s*", r"\1", str(value or "").strip().upper())
    parts = []
    for seg in code.replace("\\", "/").split("/"):
        token = re.sub(r"[^A-Z0-9+\-]", "", seg)
        if token:
            parts.append(token)
    return "/".join(parts)


def _file_sha1(path: Path) -> str:
    h = hashlib.sha1()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def collect_images() -> list[ImageInfo]:
    items: list[ImageInfo] = []
    for path in sorted(IMAGES_DIR.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED:
            continue
        items.append(
            ImageInfo(
                rel=_safe_rel(path),
                path=path,
                size_bytes=path.stat().st_size,
                binary_key=_file_sha1(path),
            )
        )
    return items


def choose_keeper(group: list[ImageInfo]) -> ImageInfo:
    def score(item: ImageInfo) -> tuple[int, int, int]:
        name = item.path.name
        code_like = 1 if re.fullmatch(r"[A-Za-z0-9+\-/]+\.(png|jpg|jpeg)", name) else 0
        return (code_like, item.size_bytes, -len(name))

    return sorted(group, key=score, reverse=True)[0]


def build_duplicate_map(images: list[ImageInfo]) -> dict[str, str]:
    groups: dict[str, list[ImageInfo]] = {}
    for info in images:
        groups.setdefault(info.binary_key, []).append(info)

    duplicate_map: dict[str, str] = {}
    for members in groups.values():
        if len(members) < 2:
            continue
        keeper = choose_keeper(members)
        for item in members:
            if item.rel != keeper.rel:
                duplicate_map[item.rel] = keeper.rel
    return duplicate_map


def patch_cache(cache_path: Path, duplicate_map: dict[str, str]) -> int:
    if not cache_path.exists():
        return 0
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return 0
    if not isinstance(data, list):
        return 0

    changed = 0
    for row in data:
        if not isinstance(row, dict):
            continue
        image_value = str(row.get("image") or "")
        if "/images/" not in image_value:
            continue
        rel = image_value.split("/images/", 1)[1].split("?", 1)[0]
        replacement = duplicate_map.get(rel)
        if not replacement:
            continue
        row["image"] = f"/images/{replacement}"
        row["image_file"] = replacement
        changed += 1

    if changed:
        cache_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return changed


def patch_excel(excel_path: Path, duplicate_map: dict[str, str]) -> int:
    if not excel_path.exists():
        return 0

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    image_col = idx.get("image")
    image_file_col = idx.get("image_file")
    code_col = idx.get("code")

    changed = 0
    for row in sheet.iter_rows(min_row=2):
        image_value = str(row[image_col].value or "").strip() if image_col is not None else ""
        image_file = str(row[image_file_col].value or "").strip().replace("\\", "/") if image_file_col is not None else ""
        code_value = str(row[code_col].value or "").strip() if code_col is not None else ""

        rel = ""
        if image_value and "/images/" in image_value:
            rel = image_value.split("/images/", 1)[1].split("?", 1)[0]
        elif image_file:
            rel = image_file

        replacement = duplicate_map.get(rel)
        if replacement:
            if image_col is not None:
                row[image_col].value = f"/images/{replacement}"
            if image_file_col is not None:
                row[image_file_col].value = replacement
            changed += 1
            continue

        # Prefer clean code-based naming for Aquant rows when possible.
        if code_value and image_file_col is not None and "Kohler/" not in image_file:
            desired_name = f"{_sanitize_code(code_value)}.png"
            if desired_name and desired_name != image_file:
                src = IMAGES_DIR / image_file
                dst = IMAGES_DIR / desired_name
                if src.exists() and not dst.exists():
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    src.rename(dst)
                    if image_col is not None:
                        row[image_col].value = f"/images/{desired_name}"
                    row[image_file_col].value = desired_name
                    changed += 1

    if changed:
        workbook.save(excel_path)
    workbook.close()
    return changed


def delete_duplicate_files(duplicate_map: dict[str, str]) -> int:
    removed = 0
    for rel in sorted(duplicate_map.keys()):
        path = IMAGES_DIR / rel
        if path.exists() and path.is_file():
            path.unlink(missing_ok=True)
            removed += 1
    return removed


def main() -> int:
    if not IMAGES_DIR.exists():
        print(f"images_dir_missing={IMAGES_DIR}")
        return 1

    images = collect_images()
    print(f"total_images_before={len(images)}")

    duplicate_map = build_duplicate_map(images)
    print(f"detected_duplicates={len(duplicate_map)}")

    cache_updates = 0
    for cache_path in CACHE_FILES:
        cache_updates += patch_cache(cache_path, duplicate_map)

    excel_updates = 0
    for excel_path in EXCEL_FILES:
        excel_updates += patch_excel(excel_path, duplicate_map)

    removed = delete_duplicate_files(duplicate_map)

    total_after = len(collect_images())
    print(f"cache_rows_updated={cache_updates}")
    print(f"excel_rows_updated={excel_updates}")
    print(f"duplicate_files_removed={removed}")
    print(f"total_images_after={total_after}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
