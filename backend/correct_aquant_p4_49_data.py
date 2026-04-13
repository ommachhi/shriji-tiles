from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

EXCEL_NAME = "aquant_catalog_p4_49.xlsx"

KNOWN_VARIANTS = {
    "CP",
    "BG",
    "BRG",
    "GG",
    "MB",
    "RG",
    "AB",
    "G",
    "BSS",
    "RGB",
    "RGW",
    "W",
    "WN",
    "SEMI",
    "TALL",
}

VARIANT_COLOR_MAP = {
    "CP": "Chrome",
    "BG": "Brushed Gold",
    "BRG": "Brushed Rose Gold",
    "GG": "Graphite Grey",
    "MB": "Matt Black",
    "RG": "Rose Gold",
    "AB": "Antique Bronze",
    "G": "Gold",
}

PRICED_GROUP_VARIANTS = {"BRG", "BG", "GG", "MB", "RG"}


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def parse_code(code_value: str) -> tuple[str, str]:
    text = normalize_spaces(code_value).upper()
    text = text.replace("MRP", "").strip(" -:")

    composite_match = re.search(r"(\d{3,5})\s*[-+]\s*(\d{3,5})\s*([A-Z]{1,4})?", text)
    if composite_match:
        first = composite_match.group(1)
        second = composite_match.group(2)
        # If the second number looks like a size token (for example 200), keep
        # the first code as the product base.
        base = second if int(second) >= 1000 else first
        suffix = (composite_match.group(3) or "").strip().upper()
        return base, suffix

    base_match = re.search(r"(\d{3,5})", text)
    if not base_match:
        return "", ""

    base = base_match.group(1)
    suffix = text[base_match.end() :].strip(" -:+/")
    suffix = re.sub(r"[^A-Z0-9]+", "", suffix)

    if suffix and suffix not in KNOWN_VARIANTS and len(suffix) > 5:
        suffix = suffix[:5]

    return base, suffix


def infer_variant_from_color(color: str) -> str:
    value = normalize_spaces(color).lower()
    if not value:
        return ""
    if "chrome" in value or value == "cp":
        return "CP"
    if "brushed rose gold" in value:
        return "BRG"
    if "brushed gold" in value:
        return "BG"
    if "graphite" in value:
        return "GG"
    if "matt black" in value or "matte black" in value:
        return "MB"
    if value == "gold" or "glossy gold" in value:
        return "G"
    if "antique bronze" in value:
        return "AB"
    if "rose gold" in value:
        return "RG"
    return ""


def build_code(base: str, variant: str) -> str:
    return f"{base} {variant}".strip() if variant else base


def build_image_filename(base: str, variant: str) -> str:
    return f"{base}_{variant}.png" if variant else f"{base}.png"


def mode_price(values: list[int]) -> int | None:
    clean = [v for v in values if isinstance(v, int) and v > 0]
    if not clean:
        return None
    counter = Counter(clean)
    return counter.most_common(1)[0][0]


def correct_excel(path: Path) -> dict[str, int | str]:
    workbook = load_workbook(path)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_cells]
    index = {name: idx for idx, name in enumerate(headers)}

    required = ["code", "color", "price", "image", "image_file", "name", "details"]
    missing = [key for key in required if key not in index]
    if missing:
        workbook.close()
        raise ValueError(f"Missing columns: {missing}")

    # Add correction columns if not present.
    next_col = len(headers) + 1
    for column_name in ["base_code", "variant", "is_cp"]:
        if column_name not in index:
            sheet.cell(row=1, column=next_col, value=column_name)
            index[column_name] = next_col - 1
            next_col += 1

    # First pass: normalize code, variant, color, image fields.
    group_rows: dict[str, list[int]] = defaultdict(list)
    corrected_rows = 0

    for row_num in range(2, sheet.max_row + 1):
        code = sheet.cell(row=row_num, column=index["code"] + 1).value
        color = sheet.cell(row=row_num, column=index["color"] + 1).value

        base, variant = parse_code(str(code or ""))
        if not base:
            continue

        # Keep variant strictly code-driven to avoid accidental remapping.
        # Color text is often descriptive and can misclassify base models.

        normalized_code = build_code(base, variant)
        image_file = build_image_filename(base, variant)

        if sheet.cell(row=row_num, column=index["code"] + 1).value != normalized_code:
            sheet.cell(row=row_num, column=index["code"] + 1, value=normalized_code)
            corrected_rows += 1

        if variant in VARIANT_COLOR_MAP:
            normalized_color = VARIANT_COLOR_MAP[variant]
            if normalize_spaces(sheet.cell(row=row_num, column=index["color"] + 1).value or "") != normalized_color:
                sheet.cell(row=row_num, column=index["color"] + 1, value=normalized_color)

        sheet.cell(row=row_num, column=index["image_file"] + 1, value=image_file)
        sheet.cell(row=row_num, column=index["image"] + 1, value=f"/images/{image_file}")
        sheet.cell(row=row_num, column=index["base_code"] + 1, value=base)
        sheet.cell(row=row_num, column=index["variant"] + 1, value=variant)
        sheet.cell(row=row_num, column=index["is_cp"] + 1, value=1 if variant == "CP" else 0)

        group_rows[base].append(row_num)

    # Second pass: price correction by variant group rules.
    price_updates = 0
    for base, rows in group_rows.items():
        variant_prices: dict[str, list[int]] = defaultdict(list)
        for row_num in rows:
            variant = normalize_spaces(sheet.cell(row=row_num, column=index["variant"] + 1).value or "").upper()
            raw_price = sheet.cell(row=row_num, column=index["price"] + 1).value
            try:
                value = int(float(raw_price or 0))
            except (TypeError, ValueError):
                value = 0
            if value > 0:
                variant_prices[variant].append(value)

        grouped_prices = []
        for variant in PRICED_GROUP_VARIANTS:
            grouped_prices.extend(variant_prices.get(variant, []))

        common_price = mode_price(grouped_prices)
        if common_price is None:
            continue

        for row_num in rows:
            variant = normalize_spaces(sheet.cell(row=row_num, column=index["variant"] + 1).value or "").upper()
            if variant in PRICED_GROUP_VARIANTS:
                raw_price = sheet.cell(row=row_num, column=index["price"] + 1).value
                try:
                    current = int(float(raw_price or 0))
                except (TypeError, ValueError):
                    current = 0
                if current != common_price:
                    sheet.cell(row=row_num, column=index["price"] + 1, value=common_price)
                    price_updates += 1

    saved_path = path
    try:
        workbook.save(path)
    except PermissionError:
        saved_path = path.with_name(f"{path.stem}_fixed{path.suffix}")
        workbook.save(saved_path)
    workbook.close()

    return {
        "corrected_rows": corrected_rows,
        "price_updates": price_updates,
        "groups": len(group_rows),
        "saved_path": str(saved_path),
    }


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    excel_path = base_dir / EXCEL_NAME
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    stats = correct_excel(excel_path)
    print(f"excel={excel_path.name}")
    print(f"corrected_rows={stats['corrected_rows']}")
    print(f"price_updates={stats['price_updates']}")
    print(f"groups={stats['groups']}")
    print(f"saved_path={stats['saved_path']}")


if __name__ == "__main__":
    main()
