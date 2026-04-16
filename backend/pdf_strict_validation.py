#!/usr/bin/env python3
"""
Strict PDF validation against Kohler catalog dataset.
This reads the entire PDF as source of truth and validates cache/Excel against it.
"""

import json
import re
from pathlib import Path
from collections import defaultdict
import fitz
from openpyxl import load_workbook

# Paths
BACKEND_DIR = Path(__file__).parent
KOHLER_PDF = BACKEND_DIR / "Kohler.pdf"
KOHLER_CACHE = BACKEND_DIR / "kohler_cache.json"
KOHLER_EXCEL = BACKEND_DIR / "kohler_catalog_full.xlsx"
IMAGES_DIR = BACKEND_DIR / "images" / "Kohler"

# Patterns - more precise to avoid matching random words
KOHLER_CODE_PATTERN = re.compile(
    r"\b(K-[\dA-Z]{1,5}(?:IN|NA|AS)?(?:-[\dA-Z]{1,4})?(?:-[\dA-Z0-9]{1,3})?|EX[\dA-Z]{6,}IN-[\dA-Z-]+)\b",
    re.I,
)
# Alternative pattern for codes that might be formatted differently
SKU_PATTERN = re.compile(r"SKU[:\s]+([K-][\dA-Z-]{5,}|EX[\dA-Z-]{8,})", re.I)

PRICE_PATTERN = re.compile(
    r"(?:MRP\s*[`\']?\s*₹?|₹|Rs\.?)\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{4,})(?:\.[0-9]{1,2})?",
    re.I,
)

def to_int_price(price_str):
    """Convert price string to integer."""
    if isinstance(price_str, int):
        return price_str
    if isinstance(price_str, float):
        return int(price_str)
    if not price_str:
        return None
    
    # Remove common currency symbols and text
    p = str(price_str).strip()
    p = re.sub(r"[^0-9,.]", "", p)
    
    # Remove commas (Indian format)
    p = p.replace(",", "")
    
    # Take only integer part
    if "." in p:
        p = p.split(".")[0]
    
    try:
        return int(p)
    except (ValueError, TypeError):
        return None

def normalize_code(code_str):
    """Normalize Kohler code for comparison."""
    if not code_str:
        return ""
    code = str(code_str).strip().upper()
    # Remove spaces and extra dashes
    code = re.sub(r"\s+", "", code)
    return code

def extract_pdf_data():
    """
    Extract all product data from PDF.
    Returns: {normalized_code: {"codes": list, "page": int, "prices": list, "bbox": rect, "text": str}}
    """
    if not KOHLER_PDF.exists():
        print(f"ERROR: PDF not found at {KOHLER_PDF}")
        return {}
    
    print(f"\n📖 Reading PDF: {KOHLER_PDF}")
    doc = fitz.open(KOHLER_PDF)
    pdf_data = {}
    
    # Exclude common English words that might match pattern
    exclude_words = {
        "EXACT", "EXTERNAL", "EXPOSED", "EXCEPTIONAL", "EXCLUSIVITY",
        "EXPANSION", "EXPERIENCE", "EXPERIENCES", "EXTENSION", "EXTN",
        "EXTREME", "EXITING"
    }
    
    for page_num, page in enumerate(doc, 1):
        text = page.get_text()
        
        # Find all codes on this page
        code_matches = list(KOHLER_CODE_PATTERN.finditer(text))
        
        for code_match in code_matches:
            raw_code = code_match.group(1)
            
            # Filter out non-codes
            if raw_code.upper() in exclude_words:
                continue
            
            # Only process if it looks like a real code
            if not (raw_code.startswith("K-") or raw_code.startswith("EX")):
                continue
            
            norm_code = normalize_code(raw_code)
            
            # Skip invalid codes
            if not norm_code or len(norm_code) < 5:
                continue
            
            if norm_code not in pdf_data:
                pdf_data[norm_code] = {
                    "codes": [],
                    "pages": set(),
                    "prices": [],
                    "text_context": "",
                    "images": []
                }
            
            if raw_code not in pdf_data[norm_code]["codes"]:
                pdf_data[norm_code]["codes"].append(raw_code)
            
            pdf_data[norm_code]["pages"].add(page_num)
            
            # Extract price near this code
            start_pos = max(0, code_match.start() - 500)
            end_pos = min(len(text), code_match.end() + 500)
            context = text[start_pos:end_pos]
            
            price_matches = list(PRICE_PATTERN.finditer(context))
            for pm in price_matches:
                price_str = pm.group(1).replace(",", "")
                try:
                    price = int(price_str)
                    if 100 <= price <= 5000000:  # Reasonable price range
                        if price not in pdf_data[norm_code]["prices"]:
                            pdf_data[norm_code]["prices"].append(price)
                except (ValueError, TypeError):
                    pass
            
            # Store text context
            if not pdf_data[norm_code]["text_context"]:
                pdf_data[norm_code]["text_context"] = context[:200]
            
            # Find images on this page
            images = page.get_images()
            if images:
                pdf_data[norm_code]["images"].append({
                    "page": page_num,
                    "count": len(images),
                    "image_list": images
                })
        
        if page_num % 10 == 0:
            print(f"  Processed page {page_num}/{len(doc)} ({len(pdf_data)} unique codes)")
    
    # Convert sets to lists for JSON serialization
    for code in pdf_data:
        pdf_data[code]["pages"] = sorted(list(pdf_data[code]["pages"]))
    
    doc.close()
    
    print(f"\n✅ PDF extraction complete: {len(pdf_data)} unique codes found")
    return pdf_data

def load_cache_data():
    """Load the kohler_cache.json dataset."""
    if not KOHLER_CACHE.exists():
        print(f"ERROR: Cache not found at {KOHLER_CACHE}")
        return {}
    
    print(f"\n📦 Loading cache: {KOHLER_CACHE}")
    with open(KOHLER_CACHE, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Normalize codes
    normalized = {}
    for product in data:
        code = normalize_code(product.get("code", ""))
        if code:
            if code not in normalized:
                normalized[code] = []
            normalized[code].append(product)
    
    print(f"✅ Loaded {len(data)} products ({len(normalized)} unique normalized codes)")
    return data, normalized

def load_excel_data():
    """Load the kohler_catalog_full.xlsx dataset."""
    if not KOHLER_EXCEL.exists():
        print(f"ERROR: Excel not found at {KOHLER_EXCEL}")
        return {}
    
    print(f"\n📊 Loading Excel: {KOHLER_EXCEL}")
    wb = load_workbook(KOHLER_EXCEL, data_only=True, read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("ERROR: Excel has no data")
        return {}
    
    # Parse header
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    try:
        code_idx = header.index("code")
        price_idx = header.index("price")
        name_idx = header.index("name")
    except ValueError as e:
        print(f"ERROR: Could not find required columns: {e}")
        return {}
    
    data = {}
    for row in rows[1:]:
        code = normalize_code(row[code_idx] if code_idx < len(row) else "")
        if code:
            price = to_int_price(row[price_idx] if price_idx < len(row) else None)
            name = str(row[name_idx] if name_idx < len(row) else "").strip() if name_idx < len(row) else ""
            
            if code not in data:
                data[code] = []
            data[code].append({
                "code": row[code_idx],
                "name": name,
                "price": price,
                "row_data": row
            })
    
    wb.close()
    print(f"✅ Loaded {len(data)} unique codes from Excel")
    return data

def check_image_exists(code):
    """Check if image file exists for code."""
    if not IMAGES_DIR.exists():
        return None, "images_dir_not_found"
    
    # Try different naming patterns
    patterns = [
        f"{code}.png",
        f"{code.lower()}.png",
        f"{code.replace('-', '_')}.png",
    ]
    
    for pattern in patterns:
        img_path = IMAGES_DIR / pattern
        if img_path.exists():
            file_size = img_path.stat().st_size
            # Check PNG signature
            try:
                with open(img_path, "rb") as f:
                    header = f.read(8)
                    if header[:4] == b"\x89PNG":
                        return str(img_path), "valid"
                    else:
                        return str(img_path), "invalid_png_signature"
            except Exception as e:
                return str(img_path), f"cannot_read: {str(e)}"
    
    # Debug: check what files exist for this code pattern
    matching_files = []
    try:
        for f in IMAGES_DIR.glob(f"{code}*"):
            matching_files.append(f.name)
    except:
        pass
    
    if matching_files:
        return None, f"pattern_mismatch_found: {matching_files[0]}"
    
    return None, "not_found"

def validate_dataset(pdf_data, cache_data, excel_data):
    """
    Compare PDF data with cache and Excel.
    Returns: {errors: [], warnings: [], summary: {}}
    """
    errors = []
    warnings = []
    stats = {
        "total_pdf_codes": len(pdf_data),
        "total_cache_products": len(cache_data[0]),
        "codes_in_pdf": set(pdf_data.keys()),
        "codes_in_cache": set(cache_data[1].keys()),
        "codes_in_excel": set(excel_data.keys()),
        "codes_only_in_pdf": set(),
        "codes_only_in_cache": set(),
        "codes_only_in_excel": set(),
        "price_mismatches": [],
        "image_missing": [],
        "image_corrupted": [],
        "correct_matches": 0,
        "total_checks": 0,
        "products_with_no_pdf_price": 0,
        "products_with_pdf_price": 0
    }
    
    # Find codes in each source only
    stats["codes_only_in_pdf"] = stats["codes_in_pdf"] - stats["codes_in_cache"]
    stats["codes_only_in_cache"] = stats["codes_in_cache"] - stats["codes_in_pdf"]
    stats["codes_only_in_excel"] = stats["codes_in_excel"] - stats["codes_in_pdf"]
    
    # Validate each code in PDF
    print("\n🔍 Validating products...")
    for norm_code, pdf_product in pdf_data.items():
        stats["total_checks"] += 1
        
        # Check if code exists in cache
        cache_products = cache_data[1].get(norm_code, [])
        excel_products = excel_data.get(norm_code, [])
        
        if not cache_products and not excel_products:
            # This code is ONLY in PDF - check if it's a real product code
            # Some codes might be artifacts - only report if they look valid
            if norm_code.startswith("K-") and len(norm_code) >= 8:
                errors.append({
                    "type": "product_not_found_in_dataset",
                    "code": norm_code,
                    "pdf_codes": pdf_product["codes"],
                    "pdf_pages": pdf_product["pages"],
                    "in_cache": False,
                    "in_excel": False,
                    "pdf_prices": pdf_product["prices"]
                })
            continue
        
        # Product exists in cache/excel
        if not pdf_product["prices"]:
            stats["products_with_no_pdf_price"] += 1
            warnings.append({
                "type": "no_price_in_pdf",
                "code": norm_code,
                "cache_product": cache_products[0] if cache_products else None
            })
            continue
        
        stats["products_with_pdf_price"] += 1
        
        # For each cache product with this code
        for cache_product in cache_products:
            cache_code = cache_product.get("code", "")
            cache_price = to_int_price(cache_product.get("price", 0))
            cache_image = cache_product.get("image", "")
            
            # Price validation
            pdf_price = pdf_product["prices"][0]  # Primary price from PDF
            
            if cache_price != pdf_price:
                # Price mismatch
                errors.append({
                    "type": "price_mismatch",
                    "code": cache_code,
                    "pdf_price": pdf_price,
                    "cache_price": cache_price,
                    "pdf_prices_all": pdf_product["prices"],
                    "pdf_pages": pdf_product["pages"],
                    "cache_name": cache_product.get("name", ""),
                    "difference": cache_price - pdf_price if cache_price else None
                })
            else:
                # Price matches
                stats["correct_matches"] += 1
            
            # Image validation
            if cache_image:
                img_file, img_status = check_image_exists(cache_image)
                
                if img_status == "not_found":
                    errors.append({
                        "type": "image_missing",
                        "code": cache_code,
                        "expected_image": cache_image,
                        "cache_name": cache_product.get("name", "")
                    })
                elif img_status.startswith("invalid") or img_status.startswith("cannot_read"):
                    errors.append({
                        "type": "image_corrupted",
                        "code": cache_code,
                        "image_file": img_file,
                        "img_status": img_status,
                        "cache_name": cache_product.get("name", "")
                    })
                # If valid or pattern mismatch, no error (not all images are used)
        
        if stats["total_checks"] % 100 == 0:
            print(f"  Validated {stats['total_checks']} codes... (errors so far: {len(errors)})")
    
    return {
        "errors": errors,
        "warnings": warnings,
        "stats": stats
    }

def generate_report(validation_result):
    """Generate and display validation report."""
    errors = validation_result["errors"]
    warnings = validation_result["warnings"]
    stats = validation_result["stats"]
    
    print("\n" + "="*80)
    print("📋 KOHLER CATALOG PDF VALIDATION REPORT")
    print("="*80)
    
    # Summary
    print("\n📊 SUMMARY:")
    print(f"  Total PDF codes found: {stats['total_pdf_codes']}")
    print(f"  Total cache products: {stats['total_cache_products']}")
    print(f"  Codes in PDF: {len(stats['codes_in_pdf'])}")
    print(f"  Codes in cache: {len(stats['codes_in_cache'])}")
    print(f"  Codes in Excel: {len(stats['codes_in_excel'])}")
    print(f"  Codes only in PDF: {len(stats['codes_only_in_pdf'])}")
    print(f"  Codes only in cache: {len(stats['codes_only_in_cache'])}")
    print(f"  Codes only in Excel: {len(stats['codes_only_in_excel'])}")
    print(f"\n  Correct matches: {stats['correct_matches']}")
    print(f"  Error count: {len(errors)}")
    print(f"  Warning count: {len(warnings)}")
    
    accuracy = (stats["correct_matches"] / stats["total_checks"] * 100) if stats["total_checks"] > 0 else 0
    print(f"\n  Accuracy: {accuracy:.2f}%")
    
    # Error breakdown
    if errors:
        print("\n❌ ERRORS ({0} total):".format(len(errors)))
        print("-" * 80)
        
        error_types = defaultdict(list)
        for error in errors:
            error_types[error["type"]].append(error)
        
        for error_type, error_list in sorted(error_types.items()):
            print(f"\n  {error_type.upper()} ({len(error_list)} items):")
            
            for error in error_list[:10]:  # Show first 10
                if error["type"] == "price_mismatch":
                    print(f"    • {error['code']}: PDF={error['pdf_price']}, Cache={error['cache_price']}")
                    print(f"      Name: {error['cache_name']}")
                    print(f"      PDF pages: {error['pdf_pages']}")
                
                elif error["type"] == "image_missing":
                    print(f"    • {error['code']}: Missing file '{error['expected_image']}'")
                    print(f"      Name: {error['cache_name']}")
                
                elif error["type"] == "image_corrupted":
                    print(f"    • {error['code']}: Corrupted image '{error['image_file']}' ({error['img_status']})")
                    print(f"      Name: {error['cache_name']}")
                
                elif error["type"] == "product_not_found_in_dataset":
                    print(f"    • {error['code']}: NOT in cache/Excel")
                    print(f"      PDF codes: {error['pdf_codes']}")
                    print(f"      PDF pages: {error['pdf_pages']}")
                    print(f"      PDF prices: {error['pdf_prices']}")
            
            if len(error_list) > 10:
                print(f"    ... and {len(error_list) - 10} more")
    
    # Warnings
    if warnings:
        print(f"\n⚠️  WARNINGS ({len(warnings)} total):")
        print("-" * 80)
        
        warning_types = defaultdict(list)
        for warning in warnings:
            warning_types[warning["type"]].append(warning)
        
        for warning_type, warning_list in sorted(warning_types.items()):
            print(f"  {warning_type.upper()}: {len(warning_list)} items")
    
    # Missing codes
    if stats["codes_only_in_pdf"]:
        print(f"\n🔔 CODES ONLY IN PDF ({len(stats['codes_only_in_pdf'])} codes):")
        print("  These codes appear in PDF but not in cache/Excel:")
        for code in sorted(stats["codes_only_in_pdf"])[:20]:
            print(f"    • {code}")
        if len(stats["codes_only_in_pdf"]) > 20:
            print(f"    ... and {len(stats['codes_only_in_pdf']) - 20} more")
    
    if stats["codes_only_in_cache"]:
        print(f"\n🔔 CODES ONLY IN CACHE ({len(stats['codes_only_in_cache'])} codes):")
        print("  These codes are in cache but not in PDF:")
        for code in sorted(stats["codes_only_in_cache"])[:20]:
            print(f"    • {code}")
        if len(stats["codes_only_in_cache"]) > 20:
            print(f"    ... and {len(stats['codes_only_in_cache']) - 20} more")
    
    # Final status
    print("\n" + "="*80)
    if len(errors) == 0:
        print("✅ VALIDATION RESULT: FULLY ACCURATE - 100% MATCH WITH PDF")
    else:
        print(f"❌ VALIDATION RESULT: ERRORS FOUND - {len(errors)} issues to resolve")
    print("="*80 + "\n")
    
    return {
        "error_count": len(errors),
        "warning_count": len(warnings),
        "accuracy_percent": accuracy,
        "errors": errors,
        "warnings": warnings,
        "stats": stats
    }

def main():
    """Main validation flow."""
    print("\n🚀 Starting strict PDF validation...\n")
    
    # Extract PDF data
    pdf_data = extract_pdf_data()
    
    # Load datasets
    cache_data = load_cache_data()
    excel_data = load_excel_data()
    
    if not pdf_data or not cache_data[0]:
        print("\n❌ Could not load necessary data")
        return
    
    # Validate
    validation_result = validate_dataset(pdf_data, cache_data, excel_data)
    
    # Generate report
    report = generate_report(validation_result)

    # Convert sets in stats to sorted lists for JSON serialization.
    if "stats" in report:
        for key in ["codes_in_pdf", "codes_in_cache", "codes_in_excel", "codes_only_in_pdf", "codes_only_in_cache", "codes_only_in_excel"]:
            if isinstance(report["stats"].get(key), set):
                report["stats"][key] = sorted(report["stats"][key])
    
    # Save report
    report_file = BACKEND_DIR / "kohler_pdf_strict_validation_report.json"
    with open(report_file, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)
    print(f"Report saved to: {report_file}")
    
    return report

if __name__ == "__main__":
    main()
