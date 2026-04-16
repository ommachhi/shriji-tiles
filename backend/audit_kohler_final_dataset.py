from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
CACHE_PATH = BASE_DIR / "kohler_cache.json"
EXCEL_PATH = BASE_DIR / "kohler_catalog_full.xlsx"
IMAGES_DIR = BASE_DIR / "images" / "Kohler"

REPORT_JSON_PATH = BASE_DIR / "kohler_final_dataset_audit_report.json"
REPORT_TXT_PATH = BASE_DIR / "kohler_final_dataset_audit_report.txt"
ANNOTATED_CACHE_PATH = BASE_DIR / "kohler_cache_annotated_price_source.json"

# Only codes manually verified from PDF in this session are tagged as pdf_verified.
PDF_VERIFIED_CODES = {
    "K1404INK0",
    "K1709INK0",
    "K72830INLAF",
    "K72830INLBL",
    "K97167INAF",
    "K97167INBL",
    "K97168INAF",
    "K97168INBL",
    "1352427",
}

LOW_PRICE_WARNING_MIN = 50
LOW_PRICE_WARNING_MAX = 500
SMALL_IMAGE_BYTES = 5_000


@dataclass
class Issue:
    code: str
    issue_type: str
    description: str
    suggested_fix: str
    severity: str

    def as_dict(self) -> dict[str, str]:
        return {
            "code": self.code,
            "issue_type": self.issue_type,
            "description": self.description,
            "suggested_fix": self.suggested_fix,
            "severity": self.severity,
        }


def normalize_code(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def parse_image_relative(image_value: str) -> str:
    image_value = clean_text(image_value)
    if not image_value:
        return ""

    image_value = image_value.replace("\\", "/")
    image_value = image_value.split("?", 1)[0]

    if "/images/" in image_value:
        image_value = image_value.split("/images/", 1)[1]
    image_value = image_value.lstrip("/")

    if image_value.lower().startswith("kohler/"):
        image_value = image_value[7:]

    return image_value


def load_cache_products(cache_path: Path) -> list[dict]:
    data = json.loads(cache_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("kohler_cache.json is not a list")
    return data


def load_excel_products(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header = [clean_text(c).lower() for c in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    required = {"code", "name", "price", "image"}
    missing_cols = [c for c in required if c not in idx]
    if missing_cols:
        wb.close()
        raise ValueError(f"Excel missing required columns: {missing_cols}")

    products: list[dict] = []
    for row in rows:
        products.append(
            {
                "code": clean_text(row[idx["code"]]),
                "name": clean_text(row[idx["name"]]),
                "price": row[idx["price"]],
                "image": clean_text(row[idx["image"]]),
                "details": clean_text(row[idx.get("details", -1)]) if "details" in idx else "",
            }
        )

    wb.close()
    return products


def to_int_price(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^0-9]", "", str(value))
    return int(digits) if digits else 0


def category_for_product(name: str, details: str, code: str) -> str:
    text = f"{name} {details} {code}".upper()
    if "WHIRLPOOL" in text or "BATHTUB" in text or "SOAKING TUB" in text:
        return "whirlpool_bathtub"
    if "RAINPANEL" in text or "RAINHEAD" in text:
        return "rainpanel_rainhead"
    if "TOILET" in text or "WC" in text or "ONE-PIECE" in text:
        return "toilet_sanitary"
    if "CLEANER" in text:
        return "cleaner"
    if "FAUCET" in text or "BASIN" in text or "TAP" in text or "SHOWER" in text:
        return "faucet_shower"
    if "ACCESSORIES" in text or "TOWEL" in text or "HOOK" in text:
        return "accessory"
    return "general"


def realistic_price_bounds(category: str) -> tuple[int, int]:
    bounds = {
        "whirlpool_bathtub": (20_000, 1_200_000),
        "rainpanel_rainhead": (2_000, 900_000),
        "toilet_sanitary": (3_000, 350_000),
        "cleaner": (100, 10_000),
        "faucet_shower": (500, 500_000),
        "accessory": (500, 200_000),
        "general": (500, 1_000_000),
    }
    return bounds[category]


def png_signature_ok(file_path: Path) -> bool:
    try:
        with file_path.open("rb") as f:
            sig = f.read(8)
        return sig == b"\x89PNG\r\n\x1a\n"
    except OSError:
        return False


def add_issue(issues: list[Issue], code: str, issue_type: str, description: str, suggested_fix: str, severity: str) -> None:
    issues.append(
        Issue(
            code=code or "<missing_code>",
            issue_type=issue_type,
            description=description,
            suggested_fix=suggested_fix,
            severity=severity,
        )
    )


def classify_price(price: int, category: str, price_source: str) -> tuple[str, str, list[Issue]]:
    """
    Returns: price_status, price_confidence, non-error warning issues.

    price_status values:
    - valid
    - low_price_needs_verification
    - invalid_ocr_price
    """
    warnings: list[Issue] = []

    if price in (1, 2, 3):
        return "invalid_ocr_price", "low", warnings

    if LOW_PRICE_WARNING_MIN <= price < LOW_PRICE_WARNING_MAX:
        warnings.append(
            Issue(
                code="",
                issue_type="low_price_needs_verification",
                description=f"Price {price} is in low verification band ({LOW_PRICE_WARNING_MIN}-{LOW_PRICE_WARNING_MAX - 1}).",
                suggested_fix="Verify against official price source; keep if confirmed.",
                severity="low",
            )
        )
        return "low_price_needs_verification", "low", warnings

    low_bound, high_bound = realistic_price_bounds(category)
    if price > 0 and (price < low_bound or price > high_bound):
        warnings.append(
            Issue(
                code="",
                issue_type="price_outlier_needs_verification",
                description=(
                    f"Price {price} outside expected range {low_bound}-{high_bound} for category {category}."
                ),
                suggested_fix="Manually verify category and official MRP.",
                severity="low",
            )
        )

    confidence = "high" if price_source == "pdf_verified" else "medium"
    return "valid", confidence, warnings


def audit() -> dict[str, Any]:
    cache_products = load_cache_products(CACHE_PATH)
    excel_products = load_excel_products(EXCEL_PATH)

    issues: list[Issue] = []
    warnings: list[Issue] = []
    high_priority: list[Issue] = []

    cache_by_code: dict[str, list[dict]] = defaultdict(list)
    for item in cache_products:
        cache_by_code[normalize_code(item.get("code"))].append(item)

    excel_by_code: dict[str, list[dict]] = defaultdict(list)
    for item in excel_products:
        excel_by_code[normalize_code(item.get("code"))].append(item)

    # Duplicate code checks in each final source.
    for code_norm, rows in cache_by_code.items():
        if code_norm and len(rows) > 1:
            add_issue(
                issues,
                rows[0].get("code", ""),
                "duplicate_code_cache",
                f"Code appears {len(rows)} times in kohler_cache.json",
                "Keep one canonical row per code and merge best fields.",
                "high",
            )

    for code_norm, rows in excel_by_code.items():
        if code_norm and len(rows) > 1:
            add_issue(
                issues,
                rows[0].get("code", ""),
                "duplicate_code_excel",
                f"Code appears {len(rows)} times in kohler_catalog_full.xlsx",
                "Keep one canonical row per code in Excel export.",
                "high",
            )

    # Canonical audit base: cache rows (runtime source), cross-validated with Excel.
    annotated_products: list[dict[str, Any]] = []

    for product in cache_products:
        code = clean_text(product.get("code"))
        code_norm = normalize_code(code)
        name = clean_text(product.get("name"))
        details = clean_text(product.get("details"))
        price = to_int_price(product.get("price"))
        image_value = clean_text(product.get("image"))

        price_source = "pdf_verified" if code_norm in PDF_VERIFIED_CODES else "ocr"

        category = category_for_product(name, details, code)
        price_status, price_confidence, price_warnings = classify_price(price, category, price_source)

        for warning in price_warnings:
            warning.code = code
            warnings.append(warning)

        annotated = dict(product)
        annotated["price_source"] = price_source
        annotated["price_status"] = price_status
        annotated["price_confidence"] = price_confidence
        annotated_products.append(annotated)

        # Rule 1: code exists
        if not code_norm:
            add_issue(
                issues,
                code,
                "missing_code",
                "Product code is empty.",
                "Populate code and ensure it is unique.",
                "high",
            )
            continue

        # Rule 2: name exists
        if not name:
            add_issue(
                issues,
                code,
                "missing_name",
                "Product name is empty.",
                "Update name from Kohler source sheet/catalog.",
                "high",
            )

        # Rule 3: price checks
        if price in (1, 2, 3):
            issue = Issue(
                code=code,
                issue_type="invalid_price_ocr_value",
                description=f"Price is {price}, invalid OCR artifact.",
                suggested_fix="Replace with verified MRP from source PDF/official list.",
                severity="high",
            )
            issues.append(issue)
            high_priority.append(issue)

        if price <= 0:
            issue = Issue(
                code=code,
                issue_type="invalid_price_nonpositive",
                description=f"Price {price} is non-positive.",
                suggested_fix="Set a valid positive MRP from final verified source.",
                severity="high",
            )
            issues.append(issue)
            high_priority.append(issue)

        # Rule 4: image checks
        relative_image = parse_image_relative(image_value)
        if not relative_image:
            issue = Issue(
                code=code,
                issue_type="missing_image",
                description="Image path is empty.",
                suggested_fix="Assign correct Kohler product image path.",
                severity="high",
            )
            issues.append(issue)
            high_priority.append(issue)
        else:
            image_file = Path(relative_image).name
            image_path = IMAGES_DIR / image_file

            if not image_path.exists():
                issue = Issue(
                    code=code,
                    issue_type="image_not_found",
                    description=f"Image file not found: {image_file}",
                    suggested_fix="Generate or copy the expected PNG into images/Kohler.",
                    severity="high",
                )
                issues.append(issue)
                high_priority.append(issue)
            else:
                # Code-image match by filename stem.
                image_stem_norm = normalize_code(image_path.stem)
                if image_stem_norm != code_norm:
                    add_issue(
                        issues,
                        code,
                        "mismatched_image_code",
                        f"Image filename {image_path.name} does not match product code {code}.",
                        "Link this product to its own code-based image file.",
                        "high",
                    )
                    high_priority.append(
                        Issue(
                            code=code,
                            issue_type="mismatched_image_code",
                            description=f"Image filename {image_path.name} does not match product code {code}.",
                            suggested_fix="Link this product to its own code-based image file.",
                            severity="high",
                        )
                    )

                try:
                    size_bytes = image_path.stat().st_size
                except OSError:
                    size_bytes = 0

                if size_bytes < SMALL_IMAGE_BYTES:
                    add_issue(
                        issues,
                        code,
                        "small_or_corrupt_image",
                        f"Image file {image_path.name} is too small ({size_bytes} bytes).",
                        "Re-render or replace image with a valid product preview.",
                        "high",
                    )
                    high_priority.append(
                        Issue(
                            code=code,
                            issue_type="small_or_corrupt_image",
                            description=f"Image file {image_path.name} is too small ({size_bytes} bytes).",
                            suggested_fix="Re-render or replace image with a valid product preview.",
                            severity="high",
                        )
                    )

                if not png_signature_ok(image_path):
                    add_issue(
                        issues,
                        code,
                        "invalid_image_format",
                        f"Image file {image_path.name} does not have valid PNG signature.",
                        "Regenerate image as valid PNG.",
                        "high",
                    )
                    high_priority.append(
                        Issue(
                            code=code,
                            issue_type="invalid_image_format",
                            description=f"Image file {image_path.name} does not have valid PNG signature.",
                            suggested_fix="Regenerate image as valid PNG.",
                            severity="high",
                        )
                    )

        # Rule 5: completeness
        critical_fields = {
            "code": bool(code_norm),
            "name": bool(name),
            "price": price > 0,
            "image": bool(relative_image),
        }
        if not all(critical_fields.values()):
            issue = Issue(
                code=code,
                issue_type="incomplete_data",
                description=f"Critical field completeness failed: {critical_fields}",
                suggested_fix="Fill missing critical fields and revalidate product row.",
                severity="high",
            )
            issues.append(issue)
            high_priority.append(issue)

        # Cross-validate against Excel row if available.
        excel_rows = excel_by_code.get(code_norm, [])
        if not excel_rows:
            add_issue(
                issues,
                code,
                "missing_in_excel",
                "Code exists in cache but not in Excel final dataset.",
                "Re-export Excel from corrected cache.",
                "high",
            )
            high_priority.append(
                Issue(
                    code=code,
                    issue_type="missing_in_excel",
                    description="Code exists in cache but not in Excel final dataset.",
                    suggested_fix="Re-export Excel from corrected cache.",
                    severity="high",
                )
            )
        else:
            excel_row = excel_rows[0]
            excel_price = to_int_price(excel_row.get("price"))
            if excel_price != price:
                add_issue(
                    issues,
                    code,
                    "price_mismatch_cache_vs_excel",
                    f"Cache price {price} differs from Excel price {excel_price}.",
                    "Sync Excel with cache corrected price.",
                    "high",
                )
            excel_name = clean_text(excel_row.get("name"))
            if excel_name and name and normalize_code(excel_name) != normalize_code(name):
                add_issue(
                    issues,
                    code,
                    "name_mismatch_cache_vs_excel",
                    "Name differs between cache and Excel sources.",
                    "Standardize product naming across final sources.",
                    "medium",
                )

    # Also detect codes present in Excel but missing in cache.
    for code_norm, rows in excel_by_code.items():
        if not code_norm:
            continue
        if code_norm not in cache_by_code:
            add_issue(
                issues,
                rows[0].get("code", ""),
                "missing_in_cache",
                "Code exists in Excel but not in cache final dataset.",
                "Rebuild cache from latest corrected dataset or add missing row.",
                "high",
            )
            high_priority.append(
                Issue(
                    code=rows[0].get("code", ""),
                    issue_type="missing_in_cache",
                    description="Code exists in Excel but not in cache final dataset.",
                    suggested_fix="Rebuild cache from latest corrected dataset or add missing row.",
                    severity="high",
                )
            )

    # Build summary.
    issue_counter = Counter(i.issue_type for i in issues)
    severity_counter = Counter(i.severity for i in issues)
    warning_counter = Counter(i.issue_type for i in warnings)

    unique_issue_codes = {normalize_code(i.code) for i in issues if normalize_code(i.code)}
    total_products = len(cache_products)
    valid_products = total_products - len(unique_issue_codes)
    quality_score = round((valid_products / total_products) * 100, 2) if total_products else 0.0

    high_severity_error_count = len([i for i in issues if i.severity == "high"])
    production_ready = len(issues) == 0 and quality_score >= 99.9 and high_severity_error_count == 0

    report = {
        "summary": {
            "total_products": total_products,
            "valid_products": valid_products,
            "error_count": len(issues),
            "warning_count": len(warnings),
            "data_quality_score_percent": quality_score,
            "production_ready": production_ready,
            "final_sources_used": ["kohler_cache.json", "kohler_catalog_full.xlsx"],
        },
        "error_breakdown": {
            "by_issue_type": dict(issue_counter),
            "by_severity": dict(severity_counter),
            "items": [i.as_dict() for i in issues],
        },
        "warnings": {
            "by_issue_type": dict(warning_counter),
            "items": [i.as_dict() for i in warnings],
        },
        "critical_issues": [i.as_dict() for i in high_priority],
        "final_validation_status": {
            "status": "PRODUCTION_READY" if production_ready else "NOT_PRODUCTION_READY",
            "reason": (
                "No high-severity issues and quality threshold met."
                if production_ready
                else "High-severity issues remain or quality threshold not met; dataset is not production-ready."
            ),
        },
        "metadata": {
            "low_price_warning_band": [LOW_PRICE_WARNING_MIN, LOW_PRICE_WARNING_MAX - 1],
            "small_image_bytes_threshold": SMALL_IMAGE_BYTES,
            "price_source_counts": dict(Counter(p["price_source"] for p in annotated_products)),
            "price_status_counts": dict(Counter(p["price_status"] for p in annotated_products)),
            "price_confidence_counts": dict(Counter(p["price_confidence"] for p in annotated_products)),
        },
    }

    ANNOTATED_CACHE_PATH.write_text(
        json.dumps(annotated_products, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    REPORT_JSON_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    REPORT_TXT_PATH.write_text(format_report_text(report), encoding="utf-8")

    return report


def format_report_text(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines: list[str] = []
    lines.append("KOHLER FINAL DATASET AUDIT REPORT")
    lines.append("=" * 72)
    lines.append("Sources: kohler_cache.json, kohler_catalog_full.xlsx")
    lines.append("")
    lines.append("1) Summary")
    lines.append("-" * 72)
    lines.append(f"Total products: {summary['total_products']}")
    lines.append(f"Valid products: {summary['valid_products']}")
    lines.append(f"Error count: {summary['error_count']}")
    lines.append(f"Warning count: {summary['warning_count']}")
    lines.append(f"Data quality score: {summary['data_quality_score_percent']}%")
    lines.append(f"Production ready: {summary['production_ready']}")
    lines.append("")

    lines.append("2) Error Breakdown")
    lines.append("-" * 72)
    for item in report["error_breakdown"]["items"]:
        lines.append(
            f"- Code: {item['code']} | Type: {item['issue_type']} | Severity: {item['severity']}"
        )
        lines.append(f"  Problem: {item['description']}")
        lines.append(f"  Suggested fix: {item['suggested_fix']}")

    lines.append("")
    lines.append("3) Warning Section (Non-critical)")
    lines.append("-" * 72)
    warning_items = report.get("warnings", {}).get("items", [])
    if not warning_items:
        lines.append("No warnings.")
    else:
        for item in warning_items:
            lines.append(f"- Code: {item['code']} | Type: {item['issue_type']} | Severity: {item['severity']}")
            lines.append(f"  Note: {item['description']}")
            lines.append(f"  Suggested action: {item['suggested_fix']}")

    lines.append("")
    lines.append("4) Critical Issues")
    lines.append("-" * 72)
    critical = report["critical_issues"]
    if not critical:
        lines.append("No high-priority critical issues.")
    else:
        for item in critical:
            lines.append(f"- {item['code']} | {item['issue_type']} | {item['description']}")

    lines.append("")
    lines.append("5) Final Validation Status")
    lines.append("-" * 72)
    lines.append(f"Status: {report['final_validation_status']['status']}")
    lines.append(f"Reason: {report['final_validation_status']['reason']}")

    return "\n".join(lines) + "\n"


def main() -> None:
    report = audit()
    print("Audit completed.")
    print(f"Total products: {report['summary']['total_products']}")
    print(f"Valid products: {report['summary']['valid_products']}")
    print(f"Error count: {report['summary']['error_count']}")
    print(f"Warning count: {report['summary']['warning_count']}")
    print(f"Data quality score: {report['summary']['data_quality_score_percent']}%")
    print(f"Production ready: {report['summary']['production_ready']}")
    print(f"JSON report: {REPORT_JSON_PATH}")
    print(f"Text report: {REPORT_TXT_PATH}")
    print(f"Annotated cache with price_source: {ANNOTATED_CACHE_PATH}")


if __name__ == "__main__":
    main()
