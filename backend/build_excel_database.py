#!/usr/bin/env python3
"""Build normalized Aquant Excel database from the full catalog PDF."""

from collections import Counter, defaultdict
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from extractor import DEFAULT_PDF_PATH, extract_products_from_pdf

OUTPUT_EXCEL = "aquant_catalog_full.xlsx"

KNOWN_VARIANTS = {
    "CP",
    "BG",
    "BRG",
    "GG",
    "MB",
    "RG",
    "AB",
    "G",
    "BSS",
    "RGB",
    "RGW",
    "W",
    "WN",
    "SEMI",
    "TALL",
}

VARIANT_COLOR_MAP = {
    "CP": "Chrome",
    "BG": "Brushed Gold",
    "BRG": "Brushed Rose Gold",
    "GG": "Graphite Grey",
    "MB": "Matt Black",
    "RG": "Rose Gold",
    "AB": "Antique Bronze",
    "G": "Gold",
}

COMMON_PRICE_VARIANTS = {"BRG", "BG", "GG", "MB", "RG"}

WINDOWS_FORBIDDEN_SEGMENT_CHARS = set('<>:"\\|?*')

# Some rows on the Brass Rain Shower page are OCR-fragmented during extraction.
# Keep these canonical entries so code-based search always returns product + image.
BRASS_RAIN_SHOWER_4000_SERIES = [
    ("4000", "Brass Rain Shower", 3100, "Chrome", "100 x 100 mm", 51),
    ("4001", "Brass Rain Shower", 3300, "Chrome", "150 mm", 51),
    ("4002", "Brass Rain Shower", 4100, "Chrome", "150 x 150 mm", 51),
    ("4003", "Brass Rain Shower", 5200, "Chrome", "200 mm", 51),
    ("4004", "Brass Rain Shower", 7500, "Chrome", "200 x 200 mm", 51),
    ("4005", "Brass Rain Shower", 8350, "Chrome", "250 mm", 51),
    ("4007", "Brass Rain Shower", 11500, "Chrome", "300 mm", 51),
    ("4008", "Brass Rain Shower", 13250, "Chrome", "300 x 300 mm", 51),
    ("4010", "Brass Rain Shower", 26250, "Chrome", "400 x 400 mm", 51),
]


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


COLOR_VARIANT_MAP = {
    normalize_spaces(color).lower(): variant
    for variant, color in VARIANT_COLOR_MAP.items()
}


def canonicalize_code(code_value: str) -> str:
    text = normalize_spaces(code_value).upper()
    text = re.sub(r"\s*([+/\-])\s*", r"\1", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -:")


def parse_code(code_value: str) -> tuple[str, str]:
    text = canonicalize_code(code_value)
    text = text.replace("MRP", "").strip(" -:")

    base_match = re.search(r"(\d{3,5})", text)
    if not base_match:
        return "", ""

    base = base_match.group(1)
    tail_variant = re.search(r"([A-Z]{1,5})$", text)
    suffix = (tail_variant.group(1) if tail_variant else "").strip().upper()
    if suffix and suffix not in KNOWN_VARIANTS and len(suffix) > 5:
        suffix = suffix[:5]
    return base, suffix


def infer_variant_from_color(color: str) -> str:
    normalized_color = normalize_spaces(color).lower()
    return COLOR_VARIANT_MAP.get(normalized_color, "")


def mode_price(values: list[int]) -> int | None:
    clean = [value for value in values if isinstance(value, int) and value > 0]
    if not clean:
        return None
    return Counter(clean).most_common(1)[0][0]


def build_image_filename(base: str, variant: str) -> str:
    return f"{base}{variant}.png" if variant else f"{base}.png"


def build_image_filename_from_code(code: str, base: str, variant: str) -> str:
    preferred = canonicalize_code(code)
    if not preferred:
        return build_image_filename(base, variant)

    # Preserve +, -, and / so filename tracks the product code format.
    preferred = preferred.replace("\\", "/")
    parts = []
    for raw_part in preferred.split("/"):
        part = raw_part.replace(" ", "").strip(".-")
        part = "".join("-" if char in WINDOWS_FORBIDDEN_SEGMENT_CHARS else char for char in part)
        if part:
            parts.append(part)

    safe = "/".join(parts).strip("/")
    if not safe or safe.startswith(".."):
        return build_image_filename(base, variant)
    return f"{safe}.png"


def normalize_products(products: list[dict]) -> list[dict]:
    by_code: dict[str, dict] = {}
    grouped_prices: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))

    def row_score(row: dict) -> tuple[int, int, int]:
        return (
            1 if row.get("price", 0) > 0 else 0,
            len(str(row.get("details") or "")),
            len(str(row.get("name") or "")),
        )

    for product in products:
        source_code = canonicalize_code(str(product.get("code") or ""))
        base_code, variant = parse_code(source_code)
        if not base_code:
            continue

        variant = variant.upper()
        color = normalize_spaces(str(product.get("color") or ""))
        if not variant and color:
            inferred_variant = infer_variant_from_color(color)
            if inferred_variant in KNOWN_VARIANTS:
                variant = inferred_variant

        if variant in VARIANT_COLOR_MAP:
            color = VARIANT_COLOR_MAP[variant]

        normalized_code = source_code or (f"{base_code}{variant}" if variant else base_code)

        price_raw = product.get("price", 0)
        try:
            price = int(float(price_raw or 0))
        except (TypeError, ValueError):
            price = 0

        row = {
            "source": product.get("source", "aquant"),
            "source_label": product.get("source_label", "Aquant"),
            "page_number": int(product.get("page_number", 0) or 0),
            "code": normalized_code,
            "name": str(product.get("name") or "").strip(),
            "size": str(product.get("size") or "").strip(),
            "color": color,
            "price": price,
            "details": str(product.get("details") or "").strip(),
            "base_code": base_code,
            "variant": variant,
            "is_cp": variant == "CP",
        }

        row["image_file"] = build_image_filename_from_code(normalized_code, base_code, variant)
        row["image"] = f"/images/{row['image_file']}"
        existing = by_code.get(normalized_code)
        if existing is None or row_score(row) > row_score(existing):
            by_code[normalized_code] = row

        if price > 0:
            grouped_prices[base_code][variant].append(price)

    normalized_rows = list(by_code.values())

    existing_codes = {str(row.get("code") or "").strip().upper() for row in normalized_rows}
    for code, name, price, color, size, page_number in BRASS_RAIN_SHOWER_4000_SERIES:
        if code in existing_codes:
            continue
        normalized_rows.append(
            {
                "source": "aquant",
                "source_label": "Aquant",
                "page_number": page_number,
                "code": code,
                "name": name,
                "size": size,
                "color": color,
                "price": price,
                "details": f"{name} {code} Size : {size}",
                "base_code": code,
                "variant": "",
                "is_cp": False,
                "image_file": build_image_filename_from_code(code, code, ""),
                "image": f"/images/{build_image_filename_from_code(code, code, '')}",
            }
        )
        existing_codes.add(code)

    common_price_by_base: dict[str, int] = {}
    for base_code, by_variant in grouped_prices.items():
        pooled = []
        for variant in COMMON_PRICE_VARIANTS:
            pooled.extend(by_variant.get(variant, []))
        common_price = mode_price(pooled)
        if common_price:
            common_price_by_base[base_code] = common_price

    for row in normalized_rows:
        variant = row["variant"]
        if variant in COMMON_PRICE_VARIANTS:
            shared_price = common_price_by_base.get(row["base_code"])
            if shared_price:
                row["price"] = shared_price

    normalized_rows.sort(key=lambda item: (item["base_code"], item["variant"], item["code"]))
    return normalized_rows

def autosize_columns(sheet):
    """Auto-size columns based on content."""
    for column_cells in sheet.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)

def export_catalog_to_excel(products: list[dict], output_path: Path) -> Path:
    """Export products to Excel with proper formatting."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    
    # Headers
    headers = [
        "source",
        "source_label",
        "page_number",
        "code",
        "name",
        "size",
        "color",
        "price",
        "details",
        "image",
        "image_file",
        "base_code",
        "variant",
        "is_cp",
    ]
    sheet.append(headers)
    
    # Format header
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for item in products:
        sheet.append([
            item.get("source", ""),
            item.get("source_label", ""),
            int(item.get("page_number", 0) or 0),
            item.get("code", ""),
            item.get("name", ""),
            item.get("size", ""),
            item.get("color", ""),
            float(item.get("price", 0) or 0),
            item.get("details", ""),
            item.get("image", ""),
            item.get("image_file", ""),
            item.get("base_code", ""),
            item.get("variant", ""),
            1 if item.get("is_cp") else 0,
        ])
    
    sheet.freeze_panes = "A2"
    autosize_columns(sheet)
    
    workbook.save(output_path)
    return output_path

def main():
    """Build normalized Excel database for all pages in Aquant catalog."""
    base_dir = Path(__file__).parent

    print("Extracting products from full PDF (all pages)...")
    raw_products = extract_products_from_pdf(
        pdf_path=DEFAULT_PDF_PATH,
        page_range=None,
        source_key="aquant",
        source_label="Aquant",
    )

    products = normalize_products(raw_products)
    print(f"Extracted raw rows: {len(raw_products)}")
    print(f"Normalized rows: {len(products)}")

    output_file = base_dir / OUTPUT_EXCEL
    print(f"Exporting to {output_file.name}...")
    export_catalog_to_excel(products, output_file)

    print(f"Complete: {len(products)} products exported")
    print(f"File: {output_file.name}")
    return output_file

if __name__ == "__main__":
    main()
