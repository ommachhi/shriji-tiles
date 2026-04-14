"""Export complete catalog data from Excel files to JSON for production deployment."""

import json
from pathlib import Path

from openpyxl import load_workbook

from extractor import image_relative_path


def export_excel_to_json(excel_path: Path, source_key: str, source_label: str) -> list[dict]:
    """Convert Excel catalog to list of product dictionaries."""
    products = []
    
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return products
    
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            wb.close()
            return products

        fields = [str(value).strip().lower() if value is not None else "" for value in header]
        index_map = {name: idx for idx, name in enumerate(fields) if name}

        def _value(row: tuple, key: str, default=""):
            idx = index_map.get(key)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        for row in rows:
            if not row:
                continue

            code = str(_value(row, "code", "")).strip()
            name = str(_value(row, "name", "")).strip()
            if not code:
                continue

            try:
                price = int(float(_value(row, "price", 0) or 0))
            except (TypeError, ValueError):
                price = 0

            color = str(_value(row, "color", "")).strip() or None
            size = str(_value(row, "size", "")).strip() or None
            details = str(_value(row, "details", name)).strip() or name
            variant = str(_value(row, "variant", "")).strip().upper() or None
            base_code = str(_value(row, "base_code", "")).strip() or None

            image_file = str(_value(row, "image_file", "")).strip()
            image = str(_value(row, "image", "")).strip()
            if image_file:
                image = f"/images/{image_relative_path(image_file)}"
            elif image:
                image = f"/images/{image_relative_path(image)}"
            else:
                image = ""

            raw_source = str(_value(row, "source", source_key)).strip().lower() or source_key
            raw_label = str(_value(row, "source_label", source_label)).strip() or source_label

            is_cp_raw = str(_value(row, "is_cp", "")).strip()
            is_cp = is_cp_raw in {"1", "true", "True", "yes", "YES"} or variant == "CP"
            
            product = {
                "source": raw_source,
                "source_label": raw_label,
                "code": code,
                "name": name,
                "price": price,
                "color": color,
                "size": size,
                "details": details,
                "base_code": base_code,
                "variant": variant,
                "is_cp": is_cp,
            }

            if image:
                product["image"] = image
            
            products.append(product)
        
        wb.close()
        print(f"✓ Exported {len(products)} products from {excel_path.name}")
        return products
    
    except Exception as e:
        print(f"✗ Error loading {excel_path}: {e}")
        return products


def main():
    backend_dir = Path(__file__).parent
    
    all_products = []
    
    # Export Aquant
    aquant_excel = backend_dir / "aquant_catalog_full.xlsx"
    aquant_products = export_excel_to_json(aquant_excel, "aquant", "Aquant")
    all_products.extend(aquant_products)
    
    # Export Kohler
    kohler_excel = backend_dir / "kohler_catalog_full.xlsx"
    kohler_products = export_excel_to_json(kohler_excel, "kohler", "Kohler")
    all_products.extend(kohler_products)
    
    # Save combined output
    output_path = backend_dir / "products_complete.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_products, f, indent=2, ensure_ascii=False)
    
    print(f"\n✓ Total: {len(all_products)} products exported to {output_path}")
    print(f"  - Aquant: {len(aquant_products)}")
    print(f"  - Kohler: {len(kohler_products)}")


if __name__ == "__main__":
    main()
