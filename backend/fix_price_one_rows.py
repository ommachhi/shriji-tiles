from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent

WORKBOOK_PATHS = [
    BASE_DIR / "aquant_catalog_full.xlsx",
    BASE_DIR / "dist" / "ProductCatalogBackend" / "_internal" / "aquant_catalog_full.xlsx",
    BASE_DIR.parent / "desktop" / "release" / "win-unpacked" / "resources" / "backend" / "_internal" / "aquant_catalog_full.xlsx",
]

CACHE_PATHS = [
    BASE_DIR / "catalog_cache.json",
]

REPORT_PATH = BASE_DIR / "price_fix_report.json"
BROKEN_LOW_PRICE_MAX = 10

SPECIAL_VARIANTS = {"AB", "BG", "BRG", "GG", "G", "MB", "RG", "BSS", "MG", "MI", "OG", "TCR", "W"}

VARIANT_COLOR_MAP = {
    "AB": "Antique Bronze",
    "BG": "Brushed Gold",
    "BRG": "Brushed Rose Gold",
    "BSS": "Brushed Stainless Steel Finish",
    "CP": "Chrome",
    "G": "Gold",
    "GG": "Graphite Grey",
    "MB": "Matt Black",
    "MG": "Matt Grey",
    "MI": "Matt Ivory",
    "OG": "Olive Green",
    "RG": "Rose Gold",
    "TCR": "Terracotta Red",
    "W": "White",
}

# Direct PDF-backed fixes for rows that were extracted as broken aliases/fragments.
EXACT_PRICE_MAP = {
    "000": 255000,
    "1040": 117000,
    "1110 MM": 197500,
    "1259": 115000,
    "1485 ABS": 8800,
    "1333": 156000,
    "1435 MM": 171500,
    "1438": 155000,
    "1860": 275000,
    "1850 W": 199500,
    "1870 W": 175000,
    "2604 BSS": 124500,
    "2634 AB": 124500,
    "2634 G": 124500,
    "2638 AB": 171500,
    "2638 G": 171500,
    "2639 AB": 197500,
    "2639 G": 197500,
    "340 MM": 175000,
    "350 MM": 149000,
    "365 MM": 199500,
    "380 MM": 185000,
    "500": 138500,
    "750": 165000,
    "850 MM": 119500,
    "855": 145000,
    "900 MM": 152000,
    "9245 CM": 80000,
    "9272": 475000,
    "9273": 271500,
    "9274": 325000,
    "30007 CP": 1700,
    "1434-600 MM": 9000,
    "1434-750 MM": 10500,
    "1434-900 MM": 12000,
    "1434-1200 MM": 18250,
}

# Families where the non-CP variants share one common price and CP has its own price.
SPECIAL_PRICE_BY_BASE = {
    "1334": 156000,
    "2113": 105000,
    "2114": 138500,
    "2115": 181500,
    "2116": 215000,
    "2122": 115000,
    "2123": 165000,
    "2124": 197500,
    "2568": 117000,
    "2569": 145000,
    "2599": 145000,
    "5013": 255000,
    "5107": 125000,
    "5122": 129500,
    "5123": 199000,
    "5134": 137000,
    "7009": 152000,
    "7515": 119500,
}

CP_PRICE_BY_BASE = {
    "2113": 85000,
    "2114": 106500,
    "2115": 138500,
    "2116": 181500,
    "2121": 55000,
    "2122": 91000,
    "2123": 123750,
    "2124": 150000,
    "2568": 91000,
    "2569": 110000,
    "2599": 129500,
    "5013": 185000,
    "5104": 27750,
    "5105": 32500,
    "5106": 32500,
    "5107": 82500,
    "5122": 99500,
    "5123": 149000,
    "5134": 99500,
    "5141": 29500,
}

CODE_METADATA_OVERRIDES = {
    "30007 CP": {
        "name": "Extendible Shower Hose (SS)",
        "details": "Extendible Shower Hose (SS) Size : 1.0 mtr",
        "size": "1.0 mtr",
        "color": "Chrome",
    },
    "9245 CM": {
        "name": "Carrara Marble Basin",
        "details": "Carrara Marble Basin",
        "size": "400 x 145 mm",
        "color": "Carrara Marble",
    },
    "1259": {
        "name": "Lagoon (316SS) Stainless Steel Out Door Shower Panel With Progressive Diverter",
        "details": "Lagoon (316SS) Stainless Steel Out Door Shower Panel With Progressive Diverter Includes : Head Shower, Hand Shower (250 mm) With Flexible Hose & Spout Matt Finish",
        "size": "2289 x 60 mm",
    },
    "1438": {
        "name": "Flora Brass Shower Column With Concealed Diverter Body",
        "details": "Flora Brass Shower Column With Concealed Diverter Body ABS Rain Shower Size : 325 x 187 mm ABS Multifunction Hand Shower With Flexible 1.5 mtr Hose Thermostatic Switch Button Diverter With Flow Control Hidden Spout Size : 1160 mm",
        "size": "1160 mm",
        "color": "Graphite Grey",
    },
    "1434-600 MM": {
        "name": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "details": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "size": "600 mm",
    },
    "1434-750 MM": {
        "name": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "details": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "size": "750 mm",
    },
    "1434-900 MM": {
        "name": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "details": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "size": "900 mm",
    },
    "1434-1200 MM": {
        "name": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "details": "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet",
        "size": "1200 mm",
    },
}

MANUAL_AQUANT_ENTRIES = [
    ("1151", "Wire - Knitted Hose (Heavy Grade 304 SS)", "Wire - Knitted Hose (Heavy Grade 304 SS)", "300 mm", "Chrome", 220, 62),
    ("1152", "Wire - Knitted Hose (Heavy Grade 304 SS)", "Wire - Knitted Hose (Heavy Grade 304 SS)", "450 mm", "Chrome", 240, 62),
    ("1153", "Wire - Knitted Hose (Heavy Grade 304 SS)", "Wire - Knitted Hose (Heavy Grade 304 SS)", "600 mm", "Chrome", 260, 62),
    ("60080 TL", "Line-Design Shower Channel Drain Without Base (304 SS)", "Line-Design Shower Channel Drain Without Base (304 SS)", "600 mm", "Stainless Steel", 2850, 57),
    ("75080 TL", "Line-Design Shower Channel Drain Without Base (304 SS)", "Line-Design Shower Channel Drain Without Base (304 SS)", "750 mm", "Stainless Steel", 3575, 57),
    ("90080 TL", "Line-Design Shower Channel Drain Without Base (304 SS)", "Line-Design Shower Channel Drain Without Base (304 SS)", "900 mm", "Stainless Steel", 5500, 57),
    ("120080 TL", "Line-Design Shower Channel Drain Without Base (304 SS)", "Line-Design Shower Channel Drain Without Base (304 SS)", "1200 mm", "Stainless Steel", 6700, 57),
    ("60080 TI", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "600 mm", "Stainless Steel", 3400, 57),
    ("75080 TI", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "750 mm", "Stainless Steel", 4250, 57),
    ("90080 TI", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "900 mm", "Stainless Steel", 6100, 57),
    ("120080 TI", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "Tile-Insert Shower Channel Drain Without Base (304 SS)", "1200 mm", "Stainless Steel", 7150, 57),
    ("60080 BS", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS)", "600 mm", "Stainless Steel", 4400, 57),
    ("75080 BS", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS)", "750 mm", "Stainless Steel", 5000, 57),
    ("90080 BS", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS)", "900 mm", "Stainless Steel", 6100, 57),
    ("120080 BS", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS)", "1200 mm", "Stainless Steel", 11000, 57),
    ("60080 BS CH", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS) CH", "600 mm", "Stainless Steel", 4400, 57),
    ("75080 BS CH", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS) CH", "750 mm", "Stainless Steel", 5000, 57),
    ("90080 BS CH", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS) CH", "900 mm", "Stainless Steel", 6100, 57),
    ("120080 BS CH", "Shower Channel Base (304 SS)", "Shower Channel Base (304 SS) CH", "1200 mm", "Stainless Steel", 11000, 57),
    ("1434-600 MM", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "600 mm", "Stainless Steel", 9000, 57),
    ("1434-750 MM", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "750 mm", "Stainless Steel", 10500, 57),
    ("1434-900 MM", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "900 mm", "Stainless Steel", 12000, 57),
    ("1434-1200 MM", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "Invisible Tile Insert Shower Channel Drainer (304 SS) Including Base With Centre-Outlet", "1200 mm", "Stainless Steel", 18250, 57),
]


@dataclass
class TemplateRow:
    name: str
    details: str
    size: str


def normalize_spaces(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def canonical_code(value: object) -> str:
    text = normalize_spaces(value).upper()
    text = re.sub(r"\s*([+/\-])\s*", r" \1 ", text)
    text = normalize_spaces(text)
    return text.strip(" -:")


def parse_base_variant(code: str) -> tuple[str, str]:
    normalized = canonical_code(code)
    match = re.match(r"^(\d{3,5})(?:\s+([A-Z]{1,5}))?", normalized)
    if not match:
        return "", ""
    return match.group(1), (match.group(2) or "").upper()


def numeric_price(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def infer_price_for_code(code: str) -> int | None:
    normalized = canonical_code(code)
    if normalized in EXACT_PRICE_MAP:
        return EXACT_PRICE_MAP[normalized]

    base, variant = parse_base_variant(normalized)
    if not base:
        return None

    if variant == "CP" and base in CP_PRICE_BY_BASE:
        return CP_PRICE_BY_BASE[base]

    if base in SPECIAL_PRICE_BY_BASE:
        if variant == "CP" and base in CP_PRICE_BY_BASE:
            return CP_PRICE_BY_BASE[base]
        if variant in SPECIAL_VARIANTS or "+" in normalized or not variant:
            return SPECIAL_PRICE_BY_BASE[base]

    return None


def template_score(name: str, details: str, size: str) -> tuple[int, int, int]:
    generic_name = name.lower() in {"", "product"}
    return (
        0 if generic_name else 1,
        len(details),
        len(size),
    )


def load_templates_from_sheet(sheet) -> dict[str, TemplateRow]:
    header = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: idx for idx, name in enumerate(header)}

    templates: dict[str, TemplateRow] = {}
    scores: dict[str, tuple[int, int, int]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = canonical_code(row[index["code"]] if "code" in index else "")
        base, _ = parse_base_variant(code)
        if not base:
            continue

        name = normalize_spaces(row[index["name"]] if "name" in index else "")
        details = normalize_spaces(row[index["details"]] if "details" in index else "")
        size = normalize_spaces(row[index["size"]] if "size" in index else "")
        score = template_score(name, details, size)
        if base not in scores or score > scores[base]:
            scores[base] = score
            templates[base] = TemplateRow(name=name, details=details, size=size)

    return templates


def _manual_entry_dict(entry: tuple[str, str, str, str, str, int, int]) -> dict:
    code, name, details, size, color, price, page_number = entry
    canonical = canonical_code(code)
    base, variant = parse_base_variant(canonical)
    image_file = f"{canonical.replace(' ', '').replace('\\', '/').replace('/', '/')}.png"
    return {
        "code": canonical,
        "name": name,
        "details": details,
        "size": size,
        "color": color,
        "price": int(price),
        "page_number": int(page_number),
        "base_code": base,
        "variant": variant,
        "is_cp": 1 if variant == "CP" else 0,
        "image_file": image_file,
        "image": f"/images/{image_file}",
    }


def _upsert_manual_workbook_rows(sheet, index: dict[str, int], changes: list[dict], path: Path) -> None:
    row_by_code: dict[str, int] = {}
    for row_num in range(2, sheet.max_row + 1):
        code = canonical_code(sheet.cell(row=row_num, column=index["code"] + 1).value)
        if code and code not in row_by_code:
            row_by_code[code] = row_num

    for entry in MANUAL_AQUANT_ENTRIES:
        manual = _manual_entry_dict(entry)
        code = manual["code"]
        row_num = row_by_code.get(code)
        if row_num is None:
            row_num = sheet.max_row + 1
            sheet.cell(row=row_num, column=index["code"] + 1, value=code)
            row_by_code[code] = row_num

        old_price = numeric_price(sheet.cell(row=row_num, column=index["price"] + 1).value) if "price" in index else 0
        changed = False
        for field, desired in manual.items():
            if field not in index:
                continue
            cell = sheet.cell(row=row_num, column=index[field] + 1)
            if field in {"price", "page_number", "is_cp"}:
                current = numeric_price(cell.value)
                desired_num = int(desired)
                if current == desired_num:
                    continue
                cell.value = desired_num
                changed = True
            else:
                current = normalize_spaces(cell.value)
                desired_text = normalize_spaces(desired)
                if current == desired_text:
                    continue
                cell.value = desired
                changed = True

        if changed:
            changes.append(
                {
                    "row": row_num,
                    "code": code,
                    "old_price": old_price,
                    "new_price": manual["price"],
                    "path": str(path),
                }
            )


def _upsert_manual_cache_rows(data: list[dict], changes: list[dict], path: Path) -> None:
    by_code: dict[str, dict] = {}
    for item in data:
        if not isinstance(item, dict):
            continue
        code = canonical_code(item.get("code", ""))
        if code and code not in by_code:
            by_code[code] = item

    for entry in MANUAL_AQUANT_ENTRIES:
        manual = _manual_entry_dict(entry)
        code = manual["code"]
        item = by_code.get(code)
        if item is None:
            item = {
                "source": "aquant",
                "source_label": "Aquant",
                "image": "",
                "image_bbox": None,
            }
            data.append(item)
            by_code[code] = item

        old_price = numeric_price(item.get("price"))
        changed = False
        for field, desired in manual.items():
            if field in {"price", "page_number"}:
                current = numeric_price(item.get(field))
                desired_num = int(desired)
                if current == desired_num:
                    continue
                item[field] = desired_num
                changed = True
            elif field == "is_cp":
                desired_bool = bool(desired)
                if bool(item.get(field)) == desired_bool:
                    continue
                item[field] = desired_bool
                changed = True
            else:
                current = normalize_spaces(item.get(field, ""))
                desired_text = normalize_spaces(desired)
                if current == desired_text:
                    continue
                item[field] = desired
                changed = True

        if changed:
            changes.append(
                {
                    "code": code,
                    "old_price": old_price,
                    "new_price": manual["price"],
                    "path": str(path),
                }
            )


def update_workbook(path: Path) -> list[dict]:
    workbook = load_workbook(path)
    sheet = workbook.active
    header = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: idx for idx, name in enumerate(header)}

    required = {"code", "price", "name", "details", "size", "color"}
    missing = sorted(required - set(index))
    if missing:
        workbook.close()
        raise ValueError(f"{path}: missing columns {missing}")

    templates = load_templates_from_sheet(sheet)
    changes: list[dict] = []

    for row_num in range(2, sheet.max_row + 1):
        code_cell = sheet.cell(row=row_num, column=index["code"] + 1)
        price_cell = sheet.cell(row=row_num, column=index["price"] + 1)
        code = canonical_code(code_cell.value)
        current_price = numeric_price(price_cell.value)
        if not code or current_price <= 0 or current_price > BROKEN_LOW_PRICE_MAX:
            continue

        new_price = infer_price_for_code(code)
        if new_price is None or new_price == current_price:
            continue

        base, variant = parse_base_variant(code)
        old_name = normalize_spaces(sheet.cell(row=row_num, column=index["name"] + 1).value)
        old_details = normalize_spaces(sheet.cell(row=row_num, column=index["details"] + 1).value)
        old_size = normalize_spaces(sheet.cell(row=row_num, column=index["size"] + 1).value)
        old_color = normalize_spaces(sheet.cell(row=row_num, column=index["color"] + 1).value)

        price_cell.value = new_price

        template = templates.get(base)
        if template:
            if old_name.lower() in {"", "product"} and template.name:
                sheet.cell(row=row_num, column=index["name"] + 1, value=template.name)
            if not old_details and template.details:
                sheet.cell(row=row_num, column=index["details"] + 1, value=template.details)
            if not old_size and template.size:
                sheet.cell(row=row_num, column=index["size"] + 1, value=template.size)

        if variant in VARIANT_COLOR_MAP:
            normalized_color = VARIANT_COLOR_MAP[variant]
            if old_color != normalized_color:
                sheet.cell(row=row_num, column=index["color"] + 1, value=normalized_color)

        override = CODE_METADATA_OVERRIDES.get(code)
        if override:
            for field, value in override.items():
                if field in index and value:
                    sheet.cell(row=row_num, column=index[field] + 1, value=value)

        changes.append(
            {
                "row": row_num,
                "code": code,
                "old_price": current_price,
                "new_price": new_price,
                "path": str(path),
            }
        )

    # Second pass: enforce PDF-mapped prices for known variant families
    # even if current value is not a low broken value.
    for row_num in range(2, sheet.max_row + 1):
        code_cell = sheet.cell(row=row_num, column=index["code"] + 1)
        price_cell = sheet.cell(row=row_num, column=index["price"] + 1)

        code = canonical_code(code_cell.value)
        if not code:
            continue

        base, _ = parse_base_variant(code)
        if not base or base not in SPECIAL_PRICE_BY_BASE:
            continue

        expected_price = infer_price_for_code(code)
        if expected_price is None:
            continue

        current_price = numeric_price(price_cell.value)
        if current_price == expected_price:
            continue

        price_cell.value = expected_price
        changes.append(
            {
                "row": row_num,
                "code": code,
                "old_price": current_price,
                "new_price": expected_price,
                "path": str(path),
            }
        )

    # Additional pass: enforce explicit exact code mappings for any current value.
    for row_num in range(2, sheet.max_row + 1):
        code = canonical_code(sheet.cell(row=row_num, column=index["code"] + 1).value)
        if not code or code not in EXACT_PRICE_MAP:
            continue
        price_cell = sheet.cell(row=row_num, column=index["price"] + 1)
        current_price = numeric_price(price_cell.value)
        expected_price = EXACT_PRICE_MAP[code]
        if current_price == expected_price:
            continue
        price_cell.value = expected_price
        changes.append(
            {
                "row": row_num,
                "code": code,
                "old_price": current_price,
                "new_price": expected_price,
                "path": str(path),
            }
        )

    # Third pass: apply metadata overrides irrespective of price state.
    for row_num in range(2, sheet.max_row + 1):
        code = canonical_code(sheet.cell(row=row_num, column=index["code"] + 1).value)
        override = CODE_METADATA_OVERRIDES.get(code)
        if not override:
            continue
        for field, value in override.items():
            if field not in index or not value:
                continue
            sheet.cell(row=row_num, column=index[field] + 1, value=value)

    # Fourth pass: ensure known missing products are present with exact prices.
    _upsert_manual_workbook_rows(sheet, index, changes, path)

    workbook.save(path)
    workbook.close()
    return changes


def update_cache(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"{path}: expected a list in cache")

    changes: list[dict] = []
    templates: dict[str, TemplateRow] = {}

    for item in data:
        if not isinstance(item, dict):
            continue
        code = canonical_code(item.get("code", ""))
        base, _ = parse_base_variant(code)
        if not base:
            continue
        name = normalize_spaces(item.get("name", ""))
        details = normalize_spaces(item.get("details", ""))
        size = normalize_spaces(item.get("size", ""))
        score = template_score(name, details, size)
        existing = templates.get(base)
        if existing is None or score > template_score(existing.name, existing.details, existing.size):
            templates[base] = TemplateRow(name=name, details=details, size=size)

    for item in data:
        if not isinstance(item, dict):
            continue
        current_price = numeric_price(item.get("price"))
        if current_price <= 0 or current_price > BROKEN_LOW_PRICE_MAX:
            continue

        code = canonical_code(item.get("code", ""))
        new_price = infer_price_for_code(code)
        if new_price is None or new_price == current_price:
            continue

        base, variant = parse_base_variant(code)
        item["price"] = new_price

        template = templates.get(base)
        if template:
            if normalize_spaces(item.get("name", "")).lower() in {"", "product"} and template.name:
                item["name"] = template.name
            if not normalize_spaces(item.get("details", "")) and template.details:
                item["details"] = template.details
            if not normalize_spaces(item.get("size", "")) and template.size:
                item["size"] = template.size

        if variant in VARIANT_COLOR_MAP:
            item["color"] = VARIANT_COLOR_MAP[variant]

        override = CODE_METADATA_OVERRIDES.get(code)
        if override:
            item.update(override)

        changes.append(
            {
                "code": code,
                "old_price": current_price,
                "new_price": new_price,
                "path": str(path),
            }
        )

    # Second pass: enforce known family prices regardless of existing value.
    for item in data:
        if not isinstance(item, dict):
            continue

        code = canonical_code(item.get("code", ""))
        if not code:
            continue

        base, _ = parse_base_variant(code)
        if not base or base not in SPECIAL_PRICE_BY_BASE:
            continue

        expected_price = infer_price_for_code(code)
        if expected_price is None:
            continue

        current_price = numeric_price(item.get("price"))
        if current_price == expected_price:
            continue

        item["price"] = expected_price
        changes.append(
            {
                "code": code,
                "old_price": current_price,
                "new_price": expected_price,
                "path": str(path),
            }
        )

    # Additional pass: enforce explicit exact code mappings for any current value.
    for item in data:
        if not isinstance(item, dict):
            continue
        code = canonical_code(item.get("code", ""))
        if not code or code not in EXACT_PRICE_MAP:
            continue
        current_price = numeric_price(item.get("price"))
        expected_price = EXACT_PRICE_MAP[code]
        if current_price == expected_price:
            continue
        item["price"] = expected_price
        changes.append(
            {
                "code": code,
                "old_price": current_price,
                "new_price": expected_price,
                "path": str(path),
            }
        )

    # Third pass: apply metadata overrides irrespective of price state.
    for item in data:
        if not isinstance(item, dict):
            continue
        code = canonical_code(item.get("code", ""))
        override = CODE_METADATA_OVERRIDES.get(code)
        if override:
            item.update(override)

    # Fourth pass: ensure known missing products are present with exact prices.
    _upsert_manual_cache_rows(data, changes, path)

    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return changes


def main() -> None:
    report = {
        "workbooks": [],
        "caches": [],
    }

    for workbook_path in WORKBOOK_PATHS:
        if not workbook_path.exists():
            continue
        changes = update_workbook(workbook_path)
        report["workbooks"].append(
            {
                "path": str(workbook_path),
                "updated_rows": len(changes),
                "changes": changes,
            }
        )

    for cache_path in CACHE_PATHS:
        if not cache_path.exists():
            continue
        changes = update_cache(cache_path)
        report["caches"].append(
            {
                "path": str(cache_path),
                "updated_rows": len(changes),
                "changes": changes,
            }
        )

    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    workbook_updates = sum(item["updated_rows"] for item in report["workbooks"])
    cache_updates = sum(item["updated_rows"] for item in report["caches"])
    print(f"workbook_updates={workbook_updates}")
    print(f"cache_updates={cache_updates}")
    print(f"report={REPORT_PATH}")


if __name__ == "__main__":
    main()
