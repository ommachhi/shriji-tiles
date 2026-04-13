#!/usr/bin/env python3
"""Regenerate images for page 5 products with improved rendering."""

import re
import sys
from pathlib import Path
from openpyxl import load_workbook
import fitz

# Ensure imports work
sys.path.insert(0, str(Path(__file__).parent))

from extractor import ensure_product_preview, DEFAULT_PDF_PATH, DEFAULT_IMAGES_DIR

def main():
    # Get products from existing Excel that are on page 5
    base_dir = Path(__file__).parent
    excel_files = sorted(base_dir.glob('catalog_pages_4_43_codeimg*p*.xlsx')) or sorted(base_dir.glob('catalog_pages_4_43_codeimg*.xlsx'))
    
    if not excel_files:
        print('ERROR: No Excel file found')
        return 1
    
    latest_excel = excel_files[-1]
    print(f'Using Excel: {latest_excel.name}')
    
    wb = load_workbook(latest_excel)
    ws = wb.active
    
    page_5_products = []
    pdf = fitz.open(DEFAULT_PDF_PATH)
    
    print('\nPage 5 Products:')
    for row in ws.iter_rows(min_row=2, max_row=300):
        if len(row) > 2 and row[2].value == 5:  # page_number column
            code = row[3].value
            name = row[4].value  
            price = row[7].value
            page_num = row[2].value
            
            # Store the row for processing
            page_5_products.append({
                'code': code,
                'name': name,
                'price': price,
                'page_number': page_num,
                'source': row[0].value or 'aquant',
                'source_label': row[1].value or 'Aquant',
                'size': row[5].value,
                'color': row[6].value,
                'details': row[8].value,
            })
            
            print(f'  {code}: {name}')
    
    pdf.close()
    
    # Regenerate images for page 5 products
    images_dir = Path(DEFAULT_IMAGES_DIR)
    images_dir.mkdir(parents=True, exist_ok=True)
    
    print(f'\n Regenerating {len(page_5_products)} product images...')
    
    # Extract fresh data for page 5 products from PDF
    from extractor import extract_products_from_pdf
    
    # We need to extract from the full catalog to get proper bboxes
    all_products = extract_products_from_pdf(
        pdf_path=DEFAULT_PDF_PATH,
        page_range=(4, 43),  # Full range to match Excel
        source_key='aquant',
        source_label='Aquant'
    )
    
    print(f'Extracted {len(all_products)} total products')
    
    # Find matching products and regenerate their images
    code_to_product = {}
    for prod in all_products:
        code = prod.get('code', '').strip()
        color = prod.get('color', '').strip()
        key = (code, color) if color else code
        if key not in code_to_product:
            code_to_product[key] = prod
    
    generated = 0
    failed = 0
    
    for product in page_5_products:
        code = product['code'].strip()
        color = product.get('color', '').strip()
        key = (code, color) if color else code
        
        if key in code_to_product:
            source_product = code_to_product[key]
            try:
                image_path = ensure_product_preview(
                    source_product,
                    pdf_path=DEFAULT_PDF_PATH,
                    images_dir=images_dir
                )
                
                # Rename to code-based name
                if image_path:
                    code_slug = re.sub(r'[^a-z0-9]+', '-', code.lower()).strip('-')
                    new_name = f"{code_slug}.png"
                    old_path = images_dir / Path(image_path).name
                    new_path = images_dir / new_name
                    
                    if old_path.exists():
                        old_path.rename(new_path)
                        generated += 1
                        print(f'  ✓ {code}: {new_name}')
            except Exception as e:
                failed += 1
                print(f'  ✗ {code}: {e}')
        else:
            print(f'  ? {code}: No matching product found in PDF')
    
    print(f'\nGeneration complete:')
    print(f'  Generated: {generated}')
    print(f'  Failed: {failed}')
    
    return 0 if failed == 0 else 1

if __name__ == '__main__':
    sys.exit(main())
