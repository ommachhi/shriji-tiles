from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from extractor import DEFAULT_IMAGES_DIR, DEFAULT_PDF_PATH, ensure_product_preview, extract_products_from_pdf


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Rebuild catalog Excel and images from PDF pages.")
    parser.add_argument("--start-page", type=int, default=1, help="Start page number (1-based)")
    parser.add_argument("--page-count", type=int, default=24, help="Number of pages to process")
    parser.add_argument(
        "--output",
        default="catalog_pages_4_43_codeimg.xlsx",
        help="Output Excel filename",
    )
    return parser.parse_args()


def code_slug(code: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(code).strip().lower()).strip("-")
    return slug or "unknown"


def cleanup_excels(base_dir: Path) -> int:
    deleted = 0
    for file_path in base_dir.glob("*.xlsx"):
        try:
            file_path.unlink()
            deleted += 1
        except PermissionError:
            # Keep locked workbook untouched.
            pass
    return deleted


def cleanup_images(images_dir: Path) -> int:
    images_dir.mkdir(parents=True, exist_ok=True)
    deleted = 0
    for file_path in images_dir.iterdir():
        if file_path.is_file():
            file_path.unlink()
            deleted += 1
    return deleted


def assign_code_based_images(products: list[dict], pdf_path: Path, images_dir: Path) -> tuple[int, int]:
    assigned = 0
    collisions = 0
    used_names: dict[str, int] = {}

    for product in products:
        image = ensure_product_preview(product, pdf_path=pdf_path, images_dir=images_dir)
        if not image:
            continue

        src_name = Path(str(image)).name
        src_path = images_dir / src_name
        if not src_path.exists():
            continue

        base_slug = code_slug(product.get("code", ""))
        ext = src_path.suffix or ".png"

        counter = used_names.get(base_slug, 0)
        while True:
            suffix = "" if counter == 0 else f"-{counter + 1}"
            target_name = f"{base_slug}{suffix}{ext}"
            target_path = images_dir / target_name
            if not target_path.exists() or target_path == src_path:
                break
            counter += 1

        used_names[base_slug] = counter
        if counter > 0:
            collisions += 1

        if src_path != target_path:
            src_path.rename(target_path)

        product["image"] = f"/images/{target_name}"
        assigned += 1

    return assigned, collisions


def export_excel(products: list[dict], output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"

    headers = [
        "source",
        "source_label",
        "page_number",
        "code",
        "name",
        "size",
        "color",
        "price",
        "details",
        "image",
        "image_file",
    ]
    sheet.append(headers)

    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in products:
        image = str(item.get("image") or "").strip()
        sheet.append(
            [
                item.get("source", ""),
                item.get("source_label", ""),
                int(item.get("page_number", 0) or 0),
                item.get("code", ""),
                item.get("name", ""),
                item.get("size", ""),
                item.get("color", ""),
                float(item.get("price", 0) or 0),
                item.get("details", ""),
                image,
                Path(image).name if image else "",
            ]
        )

    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 55)

    saved_path = output_path
    try:
        workbook.save(output_path)
    except PermissionError:
        saved_path = output_path.with_name(f"{output_path.stem}_p{len(products)}{output_path.suffix}")
        workbook.save(saved_path)
    return saved_path


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    pdf_path = Path(DEFAULT_PDF_PATH)
    images_dir = Path(DEFAULT_IMAGES_DIR)

    start_page = max(args.start_page, 1)
    end_page = max(start_page, start_page + max(args.page_count, 1) - 1)

    deleted_excels = cleanup_excels(base_dir)
    deleted_images = cleanup_images(images_dir)

    products = extract_products_from_pdf(
        pdf_path=pdf_path,
        page_range=(start_page, end_page),
        source_key="aquant",
        source_label="Aquant",
    )
    products.sort(key=lambda item: (str(item.get("code", "")), str(item.get("name", ""))))

    assigned, collisions = assign_code_based_images(products, pdf_path=pdf_path, images_dir=images_dir)

    # Keep this filename so existing backend Excel lookup can consume it directly.
    output_path = base_dir / args.output
    saved_path = export_excel(products, output_path)

    print(f"deleted_excel_files={deleted_excels}")
    print(f"deleted_images={deleted_images}")
    print(f"products_extracted={len(products)}")
    print(f"pages_processed={start_page}-{end_page}")
    print(f"images_assigned={assigned}")
    print(f"code_name_collisions={collisions}")
    print(f"excel_output={saved_path}")


if __name__ == "__main__":
    main()
