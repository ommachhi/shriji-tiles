from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

import fitz
from openpyxl import load_workbook

from extractor import DEFAULT_IMAGES_DIR, DEFAULT_PDF_PATH, _render_preview, normalize_code

EXCEL_NAME = "aquant_catalog_full.xlsx"
CACHE_NAME = "catalog_cache.json"
ROW_TOLERANCE = 18.0
BAND_GAP = 40.0
MAX_BAND_TEXT_GAP = 50.0

COLOR_PRIORITY = {
    "brushed rose gold": 0,
    "graphite grey": 1,
    "graphite gray": 1,
    "rose gold": 2,
    "brushed gold": 3,
    "gold": 3,
    "matt black": 4,
    "matte black": 4,
    "black": 4,
    "chrome": 5,
    "white glass": 5,
    "white": 5,
    "brushed stainless steel": 5,
    "stainless steel": 5,
    "stainless steel finish": 5,
    "ssf": 5,
}

VARIANT_PRIORITY = {
    "BRG": 0,
    "BR": 0,
    "GG": 1,
    "G": 1,
    "RG": 2,
    "R": 2,
    "BG": 3,
    "B": 3,
    "MB": 4,
    "M": 4,
    "CP": 5,
    "WG": 5,
    "SSF": 5,
    "BS": 5,
    "SS": 5,
}


def _compact(value: object) -> str:
    return "".join(character for character in str(value or "").upper() if character.isalnum())


def _rect_from_words(words: list[tuple]) -> fitz.Rect:
    return fitz.Rect(
        min(float(word[0]) for word in words),
        min(float(word[1]) for word in words),
        max(float(word[2]) for word in words),
        max(float(word[3]) for word in words),
    )


def _bbox_list(rect: fitz.Rect) -> list[float]:
    return [round(rect.x0, 2), round(rect.y0, 2), round(rect.x1, 2), round(rect.y1, 2)]


def _code_variant(code: str) -> str:
    parts = "".join(character if character.isalnum() else " " for character in str(code or "").upper()).split()
    return parts[-1] if parts else ""


def _display_priority(row: dict) -> tuple[int, str]:
    color_key = " ".join(str(row.get("color") or "").strip().lower().split())
    for label, priority in COLOR_PRIORITY.items():
        if label and label in color_key:
            return priority, normalize_code(row.get("code", ""))

    variant = _code_variant(str(row.get("code", "")))
    return VARIANT_PRIORITY.get(variant, 99), normalize_code(row.get("code", ""))


def _group_image_rows(image_rects: list[fitz.Rect]) -> list[dict]:
    rows: list[dict] = []

    for rect in sorted(image_rects, key=lambda item: (item.y0, item.x0)):
        center_y = (rect.y0 + rect.y1) / 2.0
        best_row = None
        best_gap = float("inf")

        for row in rows:
            gap = abs(center_y - row["center_y"])
            if gap <= ROW_TOLERANCE and gap < best_gap:
                best_gap = gap
                best_row = row

        if best_row is None:
            rows.append({"center_y": center_y, "rects": [rect]})
            continue

        best_row["rects"].append(rect)
        best_row["center_y"] = (
            best_row["center_y"] * (len(best_row["rects"]) - 1) + center_y
        ) / len(best_row["rects"])

    for row in rows:
        row["rects"].sort(key=lambda item: item.x0)

    rows.sort(key=lambda item: min(rect.y0 for rect in item["rects"]))
    return rows


def _group_image_bands(image_rects: list[fitz.Rect]) -> list[dict]:
    rows = _group_image_rows(image_rects)
    bands: list[dict] = []

    for row in rows:
        row_top = min(rect.y0 for rect in row["rects"])
        row_bottom = max(rect.y1 for rect in row["rects"])

        if not bands or row_top - bands[-1]["y1"] > BAND_GAP:
            bands.append({"rows": [row], "y0": row_top, "y1": row_bottom})
            continue

        bands[-1]["rows"].append(row)
        bands[-1]["y0"] = min(bands[-1]["y0"], row_top)
        bands[-1]["y1"] = max(bands[-1]["y1"], row_bottom)

    for band in bands:
        rects: list[fitz.Rect] = []
        for row in band["rows"]:
            rects.extend(row["rects"])
        band["rects"] = rects

    return bands


def _page_assets(page: fitz.Page) -> dict[str, list]:
    line_words: dict[tuple[int, int], list[tuple]] = defaultdict(list)
    for word in page.get_text("words"):
        line_words[(int(word[5]), int(word[6]))].append(word)

    lines = []
    for words in line_words.values():
        words.sort(key=lambda item: int(item[7]))
        text = " ".join(str(word[4]).strip() for word in words if str(word[4]).strip())
        if not text:
            continue
        lines.append({"text": text, "compact": _compact(text), "rect": _rect_from_words(words)})

    image_rects = []
    for block in page.get_text("dict")["blocks"]:
        if block.get("type") != 1:
            continue
        rect = fitz.Rect(block["bbox"])
        if rect.width < 20 or rect.height < 20:
            continue
        image_rects.append(rect)

    return {
        "lines": lines,
        "images": image_rects,
        "bands": _group_image_bands(image_rects),
    }


def _find_code_line(code: str, lines: list[dict]) -> dict | None:
    code_key = _compact(code)
    if not code_key:
        return None

    candidates = []
    for line in lines:
        line_key = line["compact"]
        if code_key not in line_key:
            continue
        exact = 0 if line_key == code_key else 1
        starts = 0 if line_key.startswith(code_key) else 1
        extra = abs(len(line_key) - len(code_key))
        candidates.append((exact, starts, extra, line["rect"].y0, line))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (item[0], item[1], item[2], item[3]))
    return candidates[0][-1]


def _choose_image_for_line(line_rect: fitz.Rect, image_rects: list[fitz.Rect]) -> fitz.Rect | None:
    best_rect: fitz.Rect | None = None
    best_cost = float("inf")
    line_center_x = (line_rect.x0 + line_rect.x1) / 2.0

    for image_rect in image_rects:
        image_center_x = (image_rect.x0 + image_rect.x1) / 2.0
        center_gap_x = abs(line_center_x - image_center_x)
        vertical_gap = line_rect.y0 - image_rect.y1

        if image_rect.y0 >= line_rect.y0 + 8:
            continue
        if vertical_gap > 140:
            continue

        x_overlap = max(
            0.0,
            min(line_rect.x1 + 6.0, image_rect.x1) - max(line_rect.x0 - 6.0, image_rect.x0),
        )
        overlap_ratio = x_overlap / max(1.0, min(line_rect.width + 12.0, image_rect.width))
        cost = abs(max(0.0, vertical_gap)) * 0.9 + center_gap_x * 0.55
        if overlap_ratio <= 0:
            cost += 90
        elif overlap_ratio < 0.22:
            cost += 28

        if cost < best_cost:
            best_cost = cost
            best_rect = image_rect

    return best_rect


def _direct_group_mapping(group_rows: list[dict], page_assets: dict[str, list]) -> dict[str, fitz.Rect] | None:
    assignments: dict[str, fitz.Rect] = {}

    for row in group_rows:
        line = _find_code_line(str(row["code"]), page_assets["lines"])
        if line is None:
            return None
        image_rect = _choose_image_for_line(line["rect"], page_assets["images"])
        if image_rect is None:
            return None
        assignments[str(row["code"])] = image_rect

    unique_rects = {tuple(_bbox_list(rect)) for rect in assignments.values()}
    if len(unique_rects) != len(group_rows):
        return None

    return assignments


def _band_group_mapping(group_rows: list[dict], page_assets: dict[str, list]) -> dict[str, fitz.Rect] | None:
    line_rects = []
    for row in group_rows:
        line = _find_code_line(str(row["code"]), page_assets["lines"])
        if line is None:
            return None
        line_rects.append(line["rect"])

    if not line_rects:
        return None

    group_top = min(rect.y0 for rect in line_rects)
    chosen_band = None
    best_gap = float("inf")

    for band in page_assets["bands"]:
        gap = group_top - band["y1"]
        if gap < -5 or gap > MAX_BAND_TEXT_GAP:
            continue
        if len(band["rects"]) != len(group_rows):
            continue
        if gap < best_gap:
            best_gap = gap
            chosen_band = band

    if chosen_band is None:
        return None

    ordered_rows = sorted(group_rows, key=_display_priority)
    ordered_rects = chosen_band["rects"]
    return {str(row["code"]): rect for row, rect in zip(ordered_rows, ordered_rects)}


def _load_rows(
    excel_path: Path,
    include_codes: set[str] | None = None,
    include_pages: set[int] | None = None,
) -> list[dict]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    index = {str(value).strip().lower(): idx for idx, value in enumerate(header)}

    required = ["code", "image_file", "page_number", "base_code"]
    for key in required:
        if key not in index:
            workbook.close()
            raise ValueError(f"Missing required column: {key}")

    loaded_rows = []
    for row in rows:
        code = str(row[index["code"]] or "").strip().upper()
        image_file = str(row[index["image_file"]] or "").strip()
        if not code or not image_file:
            continue

        try:
            page_number = int(float(row[index["page_number"]] or 0))
        except (TypeError, ValueError):
            continue

        if include_codes and normalize_code(code) not in include_codes:
            continue
        if include_pages and page_number not in include_pages:
            continue

        loaded_rows.append(
            {
                "code": code,
                "page_number": page_number,
                "image_file": image_file,
                "base_code": str(row[index["base_code"]] or "").strip() or code.split()[0],
                "color": str(row[index.get("color", -1)] or "").strip() if "color" in index else "",
            }
        )

    workbook.close()
    return loaded_rows


def _repairable_groups(group_rows: list[dict], cache_by_code: dict[str, dict], images_dir: Path) -> bool:
    if len(group_rows) < 2:
        return not (images_dir / Path(group_rows[0]["image_file"]).name).exists()

    missing_file = False
    bbox_values = []
    for row in group_rows:
        destination = images_dir / Path(str(row["image_file"]).replace("\\", "/").strip("/"))
        if not destination.exists():
            missing_file = True
        cache_item = cache_by_code.get(normalize_code(row["code"]))
        bbox = tuple(cache_item.get("image_bbox") or []) if cache_item else ()
        if bbox:
            bbox_values.append(bbox)

    if missing_file:
        return True

    if not bbox_values:
        return True

    return len(set(bbox_values)) < len(group_rows)


def regenerate_images_from_excel(
    excel_path: Path,
    images_dir: Path,
    cache_path: Path,
    include_codes: set[str] | None = None,
    include_pages: set[int] | None = None,
) -> dict[str, int]:
    loaded_rows = _load_rows(excel_path, include_codes=include_codes, include_pages=include_pages)
    cache_items = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else []
    cache_by_code = {
        normalize_code(item.get("code", "")): item
        for item in cache_items
        if isinstance(item, dict)
    }

    images_dir.mkdir(parents=True, exist_ok=True)
    grouped_rows: dict[tuple[int, str], list[dict]] = defaultdict(list)
    for row in loaded_rows:
        grouped_rows[(int(row["page_number"]), str(row["base_code"]))].append(row)

    regenerated = 0
    skipped = 0
    updated_cache = 0
    repaired_groups = 0
    page_cache: dict[int, dict[str, list]] = {}

    with fitz.open(DEFAULT_PDF_PATH) as document:
        for (page_number, _base_code), group_rows in sorted(grouped_rows.items()):
            if page_number < 0 or page_number >= len(document):
                skipped += len(group_rows)
                continue

            if not _repairable_groups(group_rows, cache_by_code, images_dir):
                continue

            page_assets = page_cache.get(page_number)
            if page_assets is None:
                page_assets = _page_assets(document.load_page(page_number))
                page_cache[page_number] = page_assets

            assignments = _direct_group_mapping(group_rows, page_assets)
            strategy = "direct"
            if assignments is None:
                assignments = _band_group_mapping(group_rows, page_assets)
                strategy = "band"

            if assignments is None:
                for row in group_rows:
                    code_key = normalize_code(row["code"])
                    cache_item = cache_by_code.get(code_key)
                    bbox = cache_item.get("image_bbox") if cache_item else None
                    if not bbox:
                        skipped += 1
                        continue
                    assignments = assignments or {}
                    assignments[row["code"]] = fitz.Rect(bbox)
                strategy = "cache"

            if not assignments:
                skipped += len(group_rows)
                continue

            repaired_groups += 1 if strategy != "cache" else 0
            for row in group_rows:
                image_rect = assignments.get(row["code"])
                if image_rect is None:
                    skipped += 1
                    continue

                destination = images_dir / Path(str(row["image_file"]).replace("\\", "/").strip("/"))
                try:
                    _render_preview(
                        document=document,
                        page_number=page_number,
                        image_bbox=_bbox_list(image_rect),
                        destination=destination,
                    )
                    regenerated += 1
                except Exception:
                    skipped += 1
                    continue

                cache_item = cache_by_code.get(normalize_code(row["code"]))
                if cache_item is None:
                    continue

                new_bbox = _bbox_list(image_rect)
                if cache_item.get("image_bbox") != new_bbox:
                    cache_item["image_bbox"] = new_bbox
                    cache_item["page_number"] = page_number
                    updated_cache += 1

    if cache_path.exists():
        cache_path.write_text(json.dumps(cache_items, indent=2, ensure_ascii=False), encoding="utf-8")

    return {
        "regenerated": regenerated,
        "skipped": skipped,
        "updated_cache": updated_cache,
        "repaired_groups": repaired_groups,
    }


def _parse_cli() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Repair Aquant image crops using code-aware PDF mapping.")
    parser.add_argument("--codes", nargs="*", default=[], help="Optional product codes to regenerate.")
    parser.add_argument("--pages", nargs="*", type=int, default=[], help="Optional zero-based PDF pages to regenerate.")
    return parser.parse_args()


def main() -> None:
    args = _parse_cli()
    base_dir = Path(__file__).resolve().parent
    excel_path = base_dir / EXCEL_NAME
    images_dir = Path(DEFAULT_IMAGES_DIR)
    cache_path = base_dir / CACHE_NAME

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    stats = regenerate_images_from_excel(
        excel_path=excel_path,
        images_dir=images_dir,
        cache_path=cache_path,
        include_codes={normalize_code(code) for code in args.codes} or None,
        include_pages=set(args.pages) or None,
    )
    print(f"excel={excel_path.name}")
    print(f"regenerated={stats['regenerated']}")
    print(f"skipped={stats['skipped']}")
    print(f"updated_cache={stats['updated_cache']}")
    print(f"repaired_groups={stats['repaired_groups']}")


if __name__ == "__main__":
    main()
