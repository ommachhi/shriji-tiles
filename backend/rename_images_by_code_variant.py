from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


def build_image_filename(code: str) -> str:
    """Build a stable filename like 2641.png or 2641_BRG.png from product code."""
    cleaned = str(code).strip().upper()
    tokens = re.findall(r"[A-Z0-9]+", cleaned)
    if not tokens:
        return "UNKNOWN.png"
    return f"{'_'.join(tokens)}.png"


def code_tokens(value: str) -> list[str]:
    return re.findall(r"[A-Z0-9]+", str(value).strip().upper())


def compact_code(value: str) -> str:
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def remap_images(excel_path: Path, images_dir: Path, cleanup_old: bool = False) -> tuple[int, int, int]:
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_cells]
    index = {name: idx for idx, name in enumerate(headers)}

    code_col = index.get("code")
    image_col = index.get("image")
    image_file_col = index.get("image_file")

    if code_col is None or image_col is None or image_file_col is None:
        workbook.close()
        raise ValueError(f"Missing required columns in {excel_path.name}")

    images_dir.mkdir(parents=True, exist_ok=True)

    updated_rows = 0
    copied_files = 0
    missing_source = 0
    used_targets: set[str] = set()
    original_names: set[str] = set()

    for row in sheet.iter_rows(min_row=2):
        code_value = row[code_col].value
        image_value = row[image_col].value

        code = str(code_value or "").strip()
        image = str(image_value or "").strip()
        if not code or not image:
            continue

        source_name = Path(image).name
        original_names.add(source_name)
        base_target = build_image_filename(code)
        target_name = base_target
        suffix = 2
        while target_name in used_targets:
            target_name = f"{Path(base_target).stem}_{suffix}.png"
            suffix += 1

        used_targets.add(target_name)
        target_path = images_dir / target_name

        source_path = images_dir / source_name
        if not source_path.exists() or not source_path.is_file():
            code_prefix = compact_code(code)
            fallback_candidates = sorted(images_dir.glob(f"{code_prefix}*.png")) if code_prefix else []
            if not fallback_candidates:
                tokens = code_tokens(code)
                if tokens:
                    base_prefix = tokens[0].lower()
                    fallback_candidates = sorted(images_dir.glob(f"{base_prefix}*.png"))
            if fallback_candidates:
                source_path = fallback_candidates[0]
            elif target_path.exists():
                source_path = target_path
            else:
                missing_source += 1
                continue

        if source_path.resolve() != target_path.resolve() and not target_path.exists():
            shutil.copy2(source_path, target_path)
            copied_files += 1

        row[image_col].value = f"/images/{target_name}"
        row[image_file_col].value = target_name
        updated_rows += 1

    keep_names: set[str] = set()
    if cleanup_old:
        keep_names = {str(row[image_file_col].value or "").strip() for row in sheet.iter_rows(min_row=2)}
        keep_names = {name for name in keep_names if name}

    workbook.save(excel_path)
    workbook.close()

    removed_files = 0
    if cleanup_old:
        for file_path in images_dir.glob("*.png"):
            if file_path.name in original_names and file_path.name not in keep_names:
                file_path.unlink(missing_ok=True)
                removed_files += 1

    return updated_rows, copied_files, missing_source + removed_files


def main() -> None:
    parser = argparse.ArgumentParser(description="Rename catalog images to code-based filenames")
    parser.add_argument(
        "--excel",
        default="aquant_catalog_p4_49.xlsx",
        help="Excel file to update (default: aquant_catalog_p4_49.xlsx)",
    )
    parser.add_argument(
        "--images-dir",
        default="images",
        help="Image directory containing existing extracted images (default: images)",
    )
    parser.add_argument(
        "--cleanup-old",
        action="store_true",
        help="Delete old image files not referenced after remapping",
    )
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    excel_path = base_dir / args.excel
    images_dir = base_dir / args.images_dir

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    updated_rows, copied_files, missing_or_removed = remap_images(
        excel_path=excel_path,
        images_dir=images_dir,
        cleanup_old=args.cleanup_old,
    )

    print(f"updated_rows={updated_rows}")
    print(f"copied_files={copied_files}")
    print(f"missing_or_removed={missing_or_removed}")
    print(f"excel={excel_path.name}")


if __name__ == "__main__":
    main()
