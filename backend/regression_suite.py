#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

import main

EXCEL_PATH = Path("aquant_catalog_full.xlsx")
IMAGES_DIR = Path("images")

WINDOWS_FORBIDDEN_SEGMENT_CHARS = set('<>:"\\|?*')


def normalized_image_file_from_code(code: str) -> str:
    value = re.sub(r"\s*([+/\-])\s*", r"\1", str(code or "").strip().upper())
    value = value.replace("\\", "/")

    parts: list[str] = []
    for raw_part in value.split("/"):
        part = raw_part.replace(" ", "").strip(".-")
        part = "".join("-" if char in WINDOWS_FORBIDDEN_SEGMENT_CHARS else char for char in part)
        if part:
            parts.append(part)

    safe = "/".join(parts).strip("/")
    return f"{safe}.png" if safe else ""


def split_image_url_path(image_url: str) -> str:
    image_url = str(image_url or "")
    if "/images/" not in image_url:
        return ""
    return image_url.split("/images/", 1)[1].split("?", 1)[0]


def check_excel_and_images() -> list[str]:
    failures: list[str] = []

    if not EXCEL_PATH.exists():
        return [f"Excel missing: {EXCEL_PATH}"]

    if not IMAGES_DIR.exists():
        return [f"Images directory missing: {IMAGES_DIR}"]

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return ["Excel has no rows"]

    header = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
    required = {"code", "base_code", "variant", "is_cp", "price", "color", "image_file"}
    missing = sorted(required - set(header))
    if missing:
        failures.append(f"Missing Excel columns: {missing}")
        return failures

    idx = {name: i for i, name in enumerate(header) if name}

    existing_images = {str(path.relative_to(IMAGES_DIR)).replace("\\", "/") for path in IMAGES_DIR.rglob("*.png")}
    if not existing_images:
        failures.append("No images generated")

    expected_images: set[str] = set()
    cp_rows = 0
    for row in rows[1:]:
        code = str(row[idx["code"]] or "").strip()
        image_file = str(row[idx["image_file"]] or "").strip().replace("\\", "/")
        is_cp = str(row[idx["is_cp"]] or "").strip()

        if is_cp in {"1", "true", "True", "yes", "YES"}:
            cp_rows += 1

        if not code:
            failures.append("Found row with empty code")
            continue

        expected = normalized_image_file_from_code(code)
        if image_file != expected:
            failures.append(f"image_file mismatch for code '{code}': got '{image_file}', expected '{expected}'")
            continue

        expected_images.add(image_file)
        if image_file not in existing_images:
            failures.append(f"Image missing on disk for code '{code}': {image_file}")

    if cp_rows == 0:
        failures.append("No CP rows detected in Excel")

    duplicates = len(existing_images) - len(set(existing_images))
    if duplicates > 0:
        failures.append(f"Duplicate image names detected: {duplicates}")

    print(f"excel_rows={len(rows) - 1}")
    print(f"excel_cp_rows={cp_rows}")
    print(f"expected_images={len(expected_images)}")
    print(f"generated_images={len(existing_images)}")
    return failures


def check_known_variant_pricing() -> list[str]:
    failures: list[str] = []
    if not EXCEL_PATH.exists():
        return [f"Excel missing: {EXCEL_PATH}"]

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return ["Excel has no rows"]

    header = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
    idx = {name: i for i, name in enumerate(header) if name}
    required = {"code", "price"}
    missing = sorted(required - set(idx))
    if missing:
        return [f"Missing pricing columns: {missing}"]

    prices = {
        str(row[idx["code"]] or "").strip().upper(): int(float(row[idx["price"]] or 0))
        for row in rows[1:]
        if str(row[idx["code"]] or "").strip()
    }

    expected_prices = {
        "2561 CP": 16500,
        "2561 BRG": 21950,
        "2561 BG": 21950,
        "2561 GG": 21950,
        "2561 MB": 21950,
        "2561 RG": 21950,
        "3161 CP": 6950,
        "3161 BRG": 9900,
        "3161 BG": 9900,
        "3161 GG": 9900,
        "3161 MB": 9900,
        "3161 RG": 9900,
        "3163 CP": 5750,
        "3163 BRG": 7100,
        "3163 BG": 7100,
        "3163 GG": 7100,
        "3163 MB": 7100,
        "3163 RG": 7100,
    }

    for code, expected_price in expected_prices.items():
        actual_price = prices.get(code)
        if actual_price != expected_price:
            failures.append(f"Price mismatch for '{code}': got {actual_price}, expected {expected_price}")

    return failures


def check_search_and_mapping() -> list[str]:
    failures: list[str] = []
    req = SimpleNamespace(base_url="http://127.0.0.1:8000/")

    critical_queries = [
        ("1320-750BRG", 1, False),
        ("30006/30007", 1, False),
        ("1336BRG+1333", 1, False),
        ("1961+1963AB", 1, False),
        ("1003 CP", 1, True),
        ("1451 GG", 1, False),
        ("2548BG", 1, False),
        ("1259", 1, False),
        ("8104GS", 1, False),
        ("1505RG", 1, False),
        ("1505", 1, False),
        ("1507", 1, False),
        ("2709", 1, False),
        ("2741", 1, False),
        ("1441", 1, False),
        ("1442", 1, False),
        ("1017", 1, False),
        ("1125", 1, False),
        ("1837", 1, False),
        ("7076", 1, False),
        ("7077", 1, False),
        ("7083", 1, False),
        ("9244", 1, False),
        ("9245", 1, False),
        ("2592 BG", 1, False),
    ]

    color_tokens = {
        "BRG": "BRUSHED ROSE",
        "BG": "BRUSHED GOLD",
        "GG": "GRAPHITE GREY",
        "MB": "MATT BLACK",
        "CP": "CHROME",
        "RG": "ROSE GOLD",
    }

    for query, min_count, require_cp in critical_queries:
        payload = main.search(req, q=query, catalog="aquant")
        rows = payload.get("results", [])
        if len(rows) < min_count:
            failures.append(f"Query '{query}' returned {len(rows)} rows, expected at least {min_count}")
            continue

        if "+" in query and len(rows) != 1:
            failures.append(f"Combined query '{query}' returned {len(rows)} rows, expected exactly 1")

        if require_cp and not any(bool(row.get("isCp")) for row in rows):
            failures.append(f"Query '{query}' should include CP product but none marked isCp")

        for row in rows[:5]:
            code = str(row.get("code") or "")
            image_url = str(row.get("image") or "")
            actual = split_image_url_path(image_url)
            expected = normalized_image_file_from_code(code)
            if actual != expected:
                failures.append(f"Search mapping mismatch for '{code}': image='{actual}', expected='{expected}'")

            variant = str(row.get("variant") or "").upper()
            color = str(row.get("color") or "").upper()
            token = color_tokens.get(variant)
            if token and token not in color:
                failures.append(f"Color mismatch for '{code}': variant={variant}, color='{row.get('color')}'")

    return failures


def main_regression() -> int:
    print("=" * 68)
    print("CATALOG REGRESSION SUITE")
    print("=" * 68)

    failures: list[str] = []
    failures.extend(check_excel_and_images())
    failures.extend(check_known_variant_pricing())
    failures.extend(check_search_and_mapping())

    print("-" * 68)
    if failures:
        print("REGRESSION STATUS: FAIL")
        for item in failures:
            print(f"FAIL: {item}")
        print(f"total_failures={len(failures)}")
        return 1

    print("REGRESSION STATUS: PASS")
    print("total_failures=0")
    return 0


if __name__ == "__main__":
    raise SystemExit(main_regression())
