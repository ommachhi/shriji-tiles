from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from extractor import (
    DEFAULT_IMAGES_DIR,
    ensure_product_preview,
    extract_products_from_pdf,
    image_relative_path,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract catalog products from PDF pages and export to Excel."
    )
    parser.add_argument(
        "--pdf",
        required=True,
        help="Path to source catalog PDF file.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to output .xlsx file.",
    )
    parser.add_argument(
        "--source-key",
        default="aquant",
        help="Internal source key (e.g. aquant, kohler).",
    )
    parser.add_argument(
        "--source-label",
        default="Aquant",
        help="Display source label.",
    )
    parser.add_argument(
        "--start-page",
        type=int,
        default=4,
        help="Start page number (1-based).",
    )
    parser.add_argument(
        "--page-count",
        type=int,
        default=40,
        help="How many pages to process.",
    )
    parser.add_argument(
        "--images-dir",
        default=str(DEFAULT_IMAGES_DIR),
        help="Image output directory for extracted previews.",
    )
    parser.add_argument(
        "--ensure-images",
        action="store_true",
        help="Generate missing preview image files (slower on large ranges).",
    )
    return parser.parse_args()


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)


def export_to_excel(products: list[dict], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"

    headers = [
        "source",
        "source_label",
        "page_number",
        "code",
        "name",
        "size",
        "color",
        "price",
        "details",
        "image",
        "image_file",
    ]
    sheet.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for product in products:
        image_path = product.get("image") or ""
        sheet.append(
            [
                product.get("source", ""),
                product.get("source_label", ""),
                int(product.get("page_number", 0) or 0),
                product.get("code", ""),
                product.get("name", ""),
                product.get("size", ""),
                product.get("color", ""),
                float(product.get("price", 0) or 0),
                product.get("details", ""),
                image_path,
                image_relative_path(image_path) if image_path else "",
            ]
        )

    sheet.freeze_panes = "A2"
    autosize_columns(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()

    pdf_path = Path(args.pdf).resolve()
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    start_page = max(args.start_page, 1)
    end_page = max(start_page, start_page + max(args.page_count, 1) - 1)

    products = extract_products_from_pdf(
        pdf_path=pdf_path,
        page_range=(start_page, end_page),
        source_key=args.source_key,
        source_label=args.source_label,
    )

    if args.ensure_images:
        images_dir = Path(args.images_dir)
        for product in products:
            image = ensure_product_preview(
                product,
                pdf_path=pdf_path,
                images_dir=images_dir,
            )
            if image:
                product["image"] = image

    products.sort(key=lambda item: (str(item.get("code", "")), str(item.get("name", ""))))

    output_path = Path(args.output).resolve()
    export_to_excel(products, output_path)

    print(f"Exported {len(products)} products")
    print(f"Pages processed: {start_page}-{end_page}")
    print(f"Excel file: {output_path}")


if __name__ == "__main__":
    main()
